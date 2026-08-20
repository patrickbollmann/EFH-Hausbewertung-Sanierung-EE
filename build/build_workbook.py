# -*- coding: utf-8 -*-
"""
Baut Hausbewertung.xlsx gemäß Bauplan (Schritt 3).
Ausführen: python3 build_workbook.py
Ausgabepfad überschreiben (z.B. für build_demo.py): Umgebungsvariable HAUSMODELL_OUTPUT setzen.
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from catalog_data import CATALOG

OUT = os.environ.get("HAUSMODELL_OUTPUT") or os.path.join(
    os.path.dirname(__file__), "..", "Hausbewertung_NRW.xlsx")

wb = Workbook()
wb.remove(wb.active)

FONT_NAME = "Arial"

# ---------- Styles ----------
def f(bold=False, size=10, color="000000", italic=False):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)

FILL_INPUT   = PatternFill("solid", fgColor="FFFF99")
FILL_HEADER  = PatternFill("solid", fgColor="1F4E5F")
FILL_SECTION = PatternFill("solid", fgColor="D9E1E2")
FILL_RESULT  = PatternFill("solid", fgColor="C6EFCE")
FILL_WARN    = PatternFill("solid", fgColor="FFC7CE")
FILL_GREY    = PatternFill("solid", fgColor="F2F2F2")

FONT_INPUT   = f(color="0000FF")
FONT_FORMULA = f(color="000000")
FONT_LINK    = f(color="006100")
FONT_RESULT  = f(bold=True, color="000000")
FONT_HEADER  = f(bold=True, color="FFFFFF")
FONT_TITLE   = f(bold=True, size=14)
FONT_SECTION = f(bold=True, size=11)
FONT_WARN    = f(bold=True, color="9C0006")
FONT_COMMENT = f(italic=True, size=8, color="666666")

THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FMT = "0.00%"
EUR_FMT = '#,##0 "€"'
EUR2_FMT = '#,##0.00 "€"'

def set_title(ws, text, cell="B1"):
    ws[cell] = text
    ws[cell].font = FONT_TITLE

def section(ws, row, text, span=6, col=2):
    c0 = get_column_letter(col)
    c1 = get_column_letter(col + span - 1)
    ws.merge_cells(f"{c0}{row}:{c1}{row}")
    cell = ws[f"{c0}{row}"]
    cell.value = text
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill = FILL_SECTION

def header_row(ws, row, headers, col=2):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col + i)
        cell.value = h
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER_ALL

def label_value(ws, row, label, value, name=None, sheet_scope=None, unit="", fmt=None,
                is_input=True, comment=None, col_label=2, col_value=3, col_unit=4, col_comment=5,
                dropdown=None, dv_list=None):
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font = f()
    vc = ws.cell(row=row, column=col_value, value=value)
    if is_input:
        vc.font = FONT_INPUT
        vc.fill = FILL_INPUT
    else:
        vc.font = FONT_FORMULA
    vc.border = BORDER_ALL
    if fmt:
        vc.number_format = fmt
    if unit:
        uc = ws.cell(row=row, column=col_unit, value=unit)
        uc.font = FONT_COMMENT
    if comment:
        cc = ws.cell(row=row, column=col_comment, value=comment)
        cc.font = FONT_COMMENT
    if name:
        ref = f"'{ws.title}'!${get_column_letter(col_value)}${row}"
        dn = DefinedName(name, attr_text=ref)
        if sheet_scope is not None:
            ws.defined_names[name] = dn
        else:
            wb.defined_names[name] = dn
    return vc

def add_dropdown(ws, cell_range, options):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv

def autofit(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# =====================================================================================
# 00_Anleitung
# =====================================================================================
ws = wb.create_sheet("00_Anleitung")
set_title(ws, "Anleitung — Excel-Modell Hausbewertung NRW")
ws["B3"] = "Farblegende"
ws["B3"].font = FONT_SECTION
legend = [
    ("Gelb, blaue Schrift", "Eingabezelle — hier tippst du Werte ein", FILL_INPUT, FONT_INPUT),
    ("Weiß, schwarze Schrift", "Formel — nicht überschreiben", None, FONT_FORMULA),
    ("Grün, fett", "Ergebnis/Kennzahl", FILL_RESULT, FONT_RESULT),
    ("Rot", "Warnung / Plausibilitätsverstoß", FILL_WARN, FONT_WARN),
    ("Kursiv grau", "Kommentar / Quellenhinweis", None, FONT_COMMENT),
]
r = 4
for label, desc, fill, font in legend:
    c = ws.cell(row=r, column=2, value=label)
    if fill: c.fill = fill
    if font: c.font = font
    ws.cell(row=r, column=3, value=desc).font = f()
    r += 1

r += 1
ws.cell(row=r, column=2, value="Neues Objekt (Haus) anlegen — Schritt für Schritt").font = FONT_SECTION
r += 1
steps = [
    "1. Objekt-ID vergeben: kurzes Kürzel ohne Unterstrich/Leerzeichen, z.B. Rahden1 oder O1. Excel erlaubt max. 31 Zeichen je Blattname, keine Sonderzeichen (/ \\ ? * [ ] :) - die ID muss kurz genug sein, dass '11_'+ID (usw.) darunter bleibt.",
    "2. Die sieben Blätter 10_VORLAGE, 11_Kalkulation, 12_Finanzierung, 13_Bauzeit, 14_Betriebskosten, 15_Zielpreis, 16_CO2_Betrieb mit Strg+Klick auf die Reiter gemeinsam markieren.",
    "3. Rechtsklick auf einen markierten Reiter -> 'Blatt verschieben/kopieren...' -> Häkchen 'Kopie erstellen' setzen -> OK. Excel kopiert alle sieben Blätter als Gruppe und passt ihre internen Querverweise automatisch an.",
    "4. Alle sieben neuen Blätter umbenennen (Doppelklick auf den Reiter): Ziffernpräfix behalten, den Rest durch die Objekt-ID ersetzen - z.B. '11_Kalkulation (2)' -> '11_Rahden1', '12_Finanzierung (2)' -> '12_Rahden1' usw. Ergebnis: 10_<ID>, 11_<ID>, 12_<ID>, 13_<ID>, 14_<ID>, 15_<ID>, 16_<ID>.",
    "5. Reiterfarbe setzen (Rechtsklick -> Registerfarbe) - eine Farbe je Objekt, damit Blattgruppen erkennbar bleiben.",
    "6. Objektdaten auf 10_<ID> eintragen (gelbe Zellen) und die Sanierungs-Checkliste durchgehen.",
    "7. Auf 00_Objektindex in der nächsten freien Zeile nur die Objekt-ID eintragen. Adresse, Kaufpreis, Wohnfläche usw. sowie die Sprunglinks erscheinen automatisch per INDIRECT()-Formel.",
    "8. 20_Dashboard übernimmt die Objekt-ID-Liste automatisch von 00_Objektindex und zieht alle Kennzahlen per INDIRECT() aus den zugehörigen 11_<ID>/12_<ID>/15_<ID>/16_<ID>-Blättern - keine manuelle Pflege, kein 'Alle aktualisieren' nötig (das Modell verwendet bewusst kein Power Query, da openpyxl keine echten Datenverbindungen erzeugen kann).",
    "9. Es ist bewusst KEIN Blattschutz aktiv (siehe Hinweis unten) - beim Ausfüllen nur die gelben Zellen "
    "anfassen, weiße Formelzellen nicht überschreiben.",
]
for s in steps:
    ws.cell(row=r, column=2, value=s).font = f()
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    r += 1

r += 1
ws.cell(row=r, column=2, value="Wichtige Hinweise").font = FONT_SECTION
r += 1
hints = [
    "DISCLAIMER (Stand 20.08.2026): Alle Förderkonditionen, -sätze und -deckel (KfW 261/358/458, BEG-EM, iSFP-Bonus usw.) sowie Zins- und Energiepreisannahmen in diesem Modell entsprechen dem Stand vom 20.08.2026 und können sich seitdem geändert haben - Förderprogramme werden erfahrungsgemäß mehrfach jährlich angepasst, ausgesetzt oder neu aufgelegt. Vor jeder Nutzung die aktuellen Konditionen auf kfw.de bzw. foerderdatenbank.de gegenprüfen (siehe auch 90_Quellen für Details je Annahme). Keine Rechts-, Steuer- oder Finanzierungsberatung.",
    "Alle Preise sind Bruttopreise Q2/2026 für die Region NRW/Ostwestfalen (siehe 90_Quellen). Vor einer Kaufentscheidung mit aktuellen Angeboten abgleichen.",
    "KfW-Zinssätze in 01_Annahmen sind Planwerte (auf kfw.de zum Recherchezeitpunkt nur Platzhalter sichtbar) - vor Finanzierungszusage bei einer Bank verifizieren.",
    "Datei enthält keine Makros (.xlsx) - läuft in Excel Windows/Mac/Online und LibreOffice.",
    "Es ist bewusst KEIN Blattschutz aktiv, damit das 'Blattgruppe kopieren' beim Anlegen neuer Objekte "
    "garantiert nicht durch Schutzeinstellungen blockiert wird. Halte dich beim Ausfüllen an die Farblegende "
    "oben (nur gelbe Zellen editieren) - Formelzellen (weiß/grün) sind bewusst ungeschützt, aber nicht zum "
    "Überschreiben gedacht.",
]
for i, h in enumerate(hints):
    cell = ws.cell(row=r, column=2, value="• " + h)
    cell.font = FONT_WARN if i == 0 else f()
    if i == 0:
        cell.fill = FILL_WARN
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    r += 1

autofit(ws, {"A": 2, "B": 14, "C": 70})
ws.sheet_view.showGridLines = False

print("00_Anleitung OK")

# =====================================================================================
# 01_Annahmen
# =====================================================================================
ws = wb.create_sheet("01_Annahmen")
set_title(ws, "01_Annahmen — globale Steuerparameter")
header_row(ws, 2, ["Parameter", "Wert", "Einheit", "Kommentar / Quelle"], col=2)
autofit(ws, {"A": 2, "B": 40, "C": 14, "D": 10, "E": 70})

r = 3

def block(title):
    global r
    section(ws, r, title, span=4)
    r += 1

def param(label, value, unit, name, comment, fmt=None, is_pct=False):
    global r
    if is_pct and fmt is None:
        fmt = PCT_FMT
    label_value(ws, r, label, value, name=name, sheet_scope=None, unit=unit,
                fmt=fmt, comment=comment)
    r += 1

block("2.1 Rahmendaten")
param("Aktuelles Jahr", 2026, "Jahr", "A_Jahr", "manuell nachführen")
param("Betrachtungshorizont Maßnahmen", 20, "Jahre", "A_Horizont", "Kreditlaufzeit-Überschneidung, deine Vorgabe")
param("Eigenkapital", 100000, "€", "A_EK", "deine Vorgabe, Default 100.000€")
param("Eigenkapital zuerst für Nebenkosten verwenden", "Ja", "Ja/Nein", "A_EK_NK_first", "Banken finanzieren Nebenkosten i.d.R. nicht")
param("Zielrabatt ggü. Neubau", 0.20, "%", "A_Zielrabatt", "deine Vorgabe", is_pct=True)
param("Zu versteuerndes Haushaltsjahreseinkommen", 60000, "€", "A_Einkommen", "steuert Einkommensbonus KfW458")
param("Minderjährige Kinder mit Kindergeld", 0, "Anzahl", "A_Kinder", "Familienzuschlag Einkommensgrenzen")
param("Objekt selbstgenutzt", "Ja", "Ja/Nein", "A_Selbstnutzung", "Voraussetzung Einkommensbonus/§35c/NRW-Programme")
param("Tragfähigkeits-Quote (Rate/Einkommen, Faustregel)", 0.35, "% Bruttoeinkommen", "A_Tragfaehigkeit_Quote",
      "grobe Faustregel fürs Dashboard-Ampel, ersetzt keine Bonitätsprüfung durch die Bank", is_pct=True)

block("2.2 Kaufnebenkosten (NRW)")
param("Grunderwerbsteuer NRW", 0.065, "%", "A_GrESt", "höchster Satz DE, kein Freibetrag Selbstnutzer 2026", is_pct=True)
param("Notar + Grundbuch", 0.018, "%", "A_Notar", "GNotKG, Spanne 1,5-2,0%", is_pct=True)
param("Maklerprovision Käuferanteil", 0.0357, "%", "A_Makler", "3,0% netto, Halbteilung §656c BGB", is_pct=True)
param("Bewegliches Inventar aus Kaufpreis", 0, "€", "A_Inventar", "mindert GrESt-Bemessungsgrundlage")

block("2.3 Preisdynamik")
param("Baupreissteigerung", 0.05, "%/a", "A_Baupreis", "Destatis Wohngebäude Mai 2026 ggü. Vorjahr", is_pct=True)
param("Baupreissteigerung Instandhaltung", 0.056, "%/a", "A_Baupreis_IH", "Destatis Instandhaltung Wohngebäude", is_pct=True)
param("Diskontsatz (Barwert)", 0.04, "%/a", "A_Diskont", "≈ Bankzins", is_pct=True)
param("Regionalfaktor OWL", 1.00, "Faktor", "A_Regio", "keine BKI-Regionaltabelle beschaffbar, neutral")
param("Risikopuffer Unvorhergesehenes", 0.12, "% Sanierungskosten", "A_Puffer", "bei Baujahr<1970 auf 15% erhöhen", is_pct=True)
param("Energiepreissteigerung", 0.03, "%/a", "A_Energiepreis", "konservative Annahme, historisch DE 3-5%/a", is_pct=True)

block("2.4 Finanzierung")
param("Bankzins Beleihung <70%", 0.0383, "% eff.", "A_Zins_60", "Interhyp 08/2026, 10J Zinsbindung", is_pct=True)
param("Bankzins Beleihung 80%", 0.0390, "% eff.", "A_Zins_80", "Interhyp 08/2026", is_pct=True)
param("Bankzins Beleihung >90%", 0.0419, "% eff.", "A_Zins_100", "Interhyp 08/2026", is_pct=True)
param("Rabatt 'grüne' Baufinanzierung", 0.0015, "%-Punkte", "A_Zins_green", "ab Effizienzklasse A/A+", is_pct=True)
param("Zinsbindung", 10, "Jahre", "A_Zinsbindung", "")
param("Anschlusszins nach Zinsbindung", 0.045, "%", "A_Zins_Anschluss", "konservative Annahme", is_pct=True)
param("Gesamtlaufzeit", 30, "Jahre", "A_Laufzeit", "")
param("Anfangstilgung Hausbank", 0.025, "%/a", "A_Tilgung", "Empfehlung ≥2%", is_pct=True)
param("KfW 261 Zinssatz", 0.030, "% eff.", "A_Zins_KfW261", "UNVERIFIZIERT - vor Antrag bei KfW-Partner prüfen", is_pct=True)
param("KfW 261 tilgungsfreie Jahre", 2, "Jahre", "A_KfW_tf", "wählbar 1-5")
param("KfW 358 Ergänzungskredit Zins", 0.025, "% eff.", "A_Zins_KfW358", "UNVERIFIZIERT", is_pct=True)
param("Bereitstellungszinsen", 0.0025, "%/Monat", "A_Bereitstellung", "entspricht 3% p.a.", is_pct=True)
param("Bereitstellungsfreie Zeit", 6, "Monate", "A_Bereit_frei", "Sanierung eher 9-12 Monate")

block("2.5 Neubau-Referenz (nur €/m² Wohnfläche)")
param("Neubau Bauwerkskosten KG300+400, GEG-Standard", 2700, "€/m² WF", "A_NB_GEG", "Schwäbisch Hall/BKI-basiert 2026")
param("Mehrkosten EH-55-Standard", 150, "€/m² WF", "A_NB_Mehrkosten_EH55", "geschätzt hälftig zu EH40, nicht separat belegt")
param("Mehrkosten EH-40-Standard", 300, "€/m² WF", "A_NB_Mehrkosten_EH40", "10-20% bzw. 200-400€/m²")
param("Baunebenkosten KG700", 0.17, "% Bauwerkskosten", "A_BNK", "15-20%", is_pct=True)
param("Außenanlagen KG500", 270, "€/m² WF", "A_Aussen", "")
param("Hausanschlüsse Neubau", 12000, "€ pauschal", "A_Anschluesse", "")

block("2.6 Energie und PV")
param("Haushaltsstrompreis", 0.30, "€/kWh", "A_P_Strom", "BDEW-Mittel 2026 37,0ct; Neuvertrag 26-30ct", is_pct=False, fmt=EUR2_FMT)
param("Wärmepumpentarif", 0.24, "€/kWh", "A_P_WPStrom", "§14a-Reduzierung", fmt=EUR2_FMT)
param("Erdgas", 0.10, "€/kWh", "A_P_Gas", "günstige Neuverträge", fmt=EUR2_FMT)
param("Heizöl", 0.139, "€/kWh", "A_P_Oel", "TECSON 20.08.2026", fmt=EUR2_FMT)
param("Pellets", 0.099, "€/kWh", "A_P_Pellet", "DEPV 08/2026", fmt=EUR2_FMT)
param("Fernwärme", 0.12, "€/kWh", "A_P_Fernwaerme", "grober Richtwert, stark anbieterabhängig - vor Ort prüfen", fmt=EUR2_FMT)
param("Einspeisevergütung ≤10kWp", 0.077, "€/kWh", "A_P_Einspeisung", "gültig 08/2026-01/2027", fmt=EUR2_FMT)
param("Spez. PV-Ertrag OWL Süd", 950, "kWh/kWp·a", "A_PV_Ertrag", "interpoliert Nord/Bund")
param("Faktor Ost-West-Ausrichtung", 0.85, "Faktor", "A_PV_OW", "80-90% des Südertrags")
param("Dachfläche je kWp", 5.5, "m²/kWp", "A_PV_Flaeche", "")
param("Nutzbarer Dachanteil", 0.50, "%", "A_PV_Anteil", "eine Seite, minus Verschattung", is_pct=True)
param("Eigenverbrauch mit WP+Speicher+EMS", 0.70, "%", "A_PV_EV", "", is_pct=True)
param("Speicher je kWp", 1.0, "kWh/kWp", "A_Speicher_kWp", "Faustformel")
param("JAZ Fußbodenheizung (35°C)", 4.5, "-", "A_JAZ_FBH", "")
param("JAZ NT-Heizkörper (45°C)", 3.8, "-", "A_JAZ_HK45", "Fraunhofer-Feldmessung Altbau Ø3,4")
param("JAZ Bestandsheizkörper (55°C)", 3.0, "-", "A_JAZ_HK55", "BEG-Mindestwert Luft/Wasser")
param("Warmwasser je Person", 1000, "kWh/a", "A_WW_Person", "berechneter Wert, nicht Pauschale 500kWh")
param("Haushaltsstrom (ohne Heizung) je Person", 1300, "kWh/a", "A_HH_Strom_Person",
      "Faustwert BDEW-Stromspiegel, grobe Näherung ohne Skaleneffekt")
param("Personen im Haushalt (Default)", 3, "Anzahl", "A_Personen", "")
param("CO2-Faktor Strommix", 344, "g/kWh", "A_CO2_Strom", "UBA 2025 real; GEG rechnet mit 560")
param("CO2-Faktor Erdgas", 201, "g/kWh", "A_CO2_Gas", "")
param("CO2-Faktor Heizöl", 266, "g/kWh", "A_CO2_Oel", "")
param("CO2-Faktor Pellets", 21, "g/kWh", "A_CO2_Pellet", "")
param("CO2-Faktor Fernwärme", 180, "g/kWh", "A_CO2_Fernwaerme", "grober Mittelwert Fernwärme-Mix DE, stark netzabhängig")
param("CO2-Gutschrift PV-Einspeisung", 344, "g/kWh", "A_CO2_PV", "Verdrängungsmix = Strommix")

block("2.7 Mengenformel-Faktoren")
param("Dachflächenfaktor Satteldach", 1.30, "×Grundfläche", "A_F_Dach", "1/cos35°+Überstand")
param("Dachflächenfaktor Flachdach", 1.05, "×Grundfläche", "A_F_Dach_flach", "")
param("Umfangfaktor", 4.10, "×√Grundfläche", "A_F_Umfang", "Rechteckform 1:1,3")
param("Geschosshöhe", 2.90, "m", "A_F_Geschoss", "")
param("Fensterflächenanteil an Fassade", 0.15, "%", "A_F_Fenster", "typisch 12-18%", is_pct=True)
param("Mittlere Fenstergröße", 1.70, "m²/Stk", "A_F_Fenstergr", "")
param("Gerüstzuschlag auf Fassadenfläche", 1.15, "Faktor", "A_F_Geruest", "")
param("Beheizte Fläche für Flächenheizung", 0.90, "% WF", "A_F_FBH", "", is_pct=True)
param("Heizkörper je m² WF", 1/15, "Stück/m²", "A_F_HK", "ein Heizkörper je 15m²")
param("Innentüren je m² WF", 1/20, "Stück/m²", "A_F_Tueren", "")
param("Räume je m² WF (für KNX)", 1/18, "Stück/m²", "A_F_Raeume", "")

block("2.8 Eigenleistung")
param("Stundensatz Eigenleistung (Bewertung)", 0, "€/h", "A_EL_Satz", "nur zur Bewertung, keine Auszahlung")
param("Materialanteil, nicht einsparbar", 0.55, "%", "A_EL_Material", "", is_pct=True)
param("Warnschwelle Eigenleistungsanteil", 0.15, "% Sanierungskosten", "A_EL_Warn", "", is_pct=True)

ws.freeze_panes = "B3"
print("01_Annahmen OK, letzte Zeile:", r)

# =====================================================================================
# 02_Massnahmenkatalog
# =====================================================================================
ws = wb.create_sheet("02_Massnahmenkatalog")
set_title(ws, "02_Massnahmenkatalog — Preise, Nutzungsdauern, Förderregeln (Q2/2026, Bruttopreise NRW/OWL)")
CAT_HEADERS = ["ID", "Kategorie", "Maßnahme", "Einheit", "Einheitspreis €", "Nutzungsdauer (Jahre)",
               "Förderfähig", "Fördersatz Grundprogramm", "Pflicht EH55EE", "Pflicht EH40EE",
               "Klimaneutralität", "Eigenleistung möglich", "Kommentar / Quelle"]
CAT_HEADER_ROW = 3
header_row(ws, CAT_HEADER_ROW, CAT_HEADERS, col=2)

CAT_FIRST_ROW = CAT_HEADER_ROW + 1
CAT_ROW = {}
for i, item in enumerate(CATALOG):
    row = CAT_FIRST_ROW + i
    CAT_ROW[item["id"]] = row
    vals = [item["id"], item["kat"], item["name"], item["einheit"], item["preis"], item["nutzung"],
            "Ja" if item["foerder"] else "Nein", item["satz"],
            "Ja" if item["eh55"] else "Nein", "Ja" if item["eh40"] else "Nein",
            "Ja" if item["klima"] else "Nein", "Ja" if item["el"] else "Nein", item["kommentar"]]
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=2 + j, value=v)
        c.font = FONT_FORMULA
        c.border = BORDER_ALL
        if j == 4:
            c.number_format = EUR_FMT
        if j == 2 or j == 12:
            c.alignment = Alignment(wrap_text=True, vertical="top")
CAT_LAST_ROW = CAT_FIRST_ROW + len(CATALOG) - 1

autofit(ws, {"A": 2, "B": 8, "C": 20, "D": 46, "E": 20, "F": 14, "G": 12,
             "H": 12, "I": 26, "J": 12, "K": 12, "L": 14, "M": 14, "N": 55})
ws.freeze_panes = "B4"
ws.auto_filter.ref = f"B{CAT_HEADER_ROW}:N{CAT_LAST_ROW}"
print("02_Massnahmenkatalog OK, Zeilen", CAT_FIRST_ROW, "-", CAT_LAST_ROW)

# =====================================================================================
# 03_Foerderprogramme
# =====================================================================================
ws = wb.create_sheet("03_Foerderprogramme")
set_title(ws, "03_Foerderprogramme — Konditionen, Höchstbeträge, Wegevergleich (Stand Merkblätter 07/2026)")
section(ws, 3, "4.1 Programmtabelle", span=6)
FP_HEADERS = ["Programm", "Träger", "Gegenstand", "Satz/Höhe", "Höchstbetrag", "Voraussetzung"]
header_row(ws, 4, FP_HEADERS, col=2)
FP_ROWS = [
    ("KfW 261", "KfW", "Effizienzhaus-Kredit",
     "Tilgungszuschuss 5% (EH55EE) / 10% (EH40EE), +10%-Pkt. WPB, +15%-Pkt. Serielle San.",
     "Kredit 150.000€/WE", "EE-Klasse ≥65% erneuerbar ist Pflichtvoraussetzung (nicht mehr Bonus)"),
    ("BAFA BEG-EM Gebäudehülle", "BAFA", "Dämmung, Fenster, Lüftung, digitale Energiesysteme",
     "15% + 5% iSFP (ab 30.000€ Investition)", "30.000€ (60.000€ mit iSFP) je WE/Jahr",
     "Energieeffizienz-Experte, Vertrag mit aufschiebender Bedingung"),
    ("BAFA Heizungsoptimierung", "BAFA", "Hydraulischer Abgleich, Pumpentausch etc.", "15%",
     "im Hüllentopf enthalten", "Mindestinvestition 300€"),
    ("KfW 458", "KfW", "Wärmepumpe/Heizungstausch",
     "30% Grund + 16% Klimageschwindigkeit + 0-40% Einkommen", "28.000€ 1. WE",
     "Austausch fossile/alte Heizung, Gesamtsatz max. 70-80%"),
    ("KfW 358/359", "KfW", "Ergänzungskredit zu BEG-EM",
     "zinsverbilligt (358, Einkommen ≤90.000€) bzw. zinsfrei-Alt. (359)", "120.000€/WE",
     "nur zusätzlich zu bestehender Zuschusszusage, Zusage <12 Monate alt"),
    ("BAFA Energieberatung/iSFP", "BAFA", "Beratungshonorar", "50%", "650€ (EFH)", ""),
    ("§35c EStG", "Finanzamt", "Steuerbonus (Alternative zu BEG)", "7+7+6% über 3 Jahre = 20%",
     "40.000€ je Objekt", "Gebäude >10 Jahre; NICHT kombinierbar mit BEG-Zuschuss/-Kredit für dieselbe Maßnahme"),
    ("NRW.BANK Modernisierung", "NRW.BANK", "Darlehen einkommensabhängig",
     "0% Zins 5J, dann 0,5%, Tilgungsnachlass 25-50%", "5.000-220.000€",
     "Einkommensgrenze Gruppe A/B"),
    ("NRW.BANK.Wohneigentum", "NRW.BANK", "Kauf-/Anschlussfinanzierung", "Marktzins, einkommensabhängig zugänglich",
     "kein Deckel", "Einkommensgrenze 75.000/100.000€ + 20.000€/Kind"),
    ("Kreis Minden-Lübbecke 'Klimaresilienz'", "Kreis", "Entsiegelung, Dach-/Fassadenbegrünung",
     "unbekannt (Richtlinie nicht abrufbar)", "unbekannt", "Mittelreservierung anfragen: 0571 807-26801"),
]
r = 5
for row_data in FP_ROWS:
    for j, v in enumerate(row_data):
        c = ws.cell(row=r, column=2 + j, value=v)
        c.font = FONT_FORMULA
        c.border = BORDER_ALL
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
section(ws, r, "4.2 Kumulierungsgrenze", span=6); r += 1
ws.cell(row=r, column=2, value="Max. 60% der Investition aus öffentlichen Mitteln (KfW-Merkblatt 261). Heizungsförderung hat eigenen Deckel 70-80% (KfW 458).").font = f()
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
FP_KUMUL_ROW = r
r += 2

section(ws, r, "4.3 Förderweg-Vergleich (je Objekt aus 11_Kalkulation verlinkt)", span=6); r += 1
FW_LABEL_ROW = r
labels = ["", "Weg A: Effizienzhaus (KfW 261)", "Weg B: Einzelmaßnahmen (BEG-EM + KfW 458)"]
for j, v in enumerate(labels):
    c = ws.cell(row=r, column=2 + j, value=v)
    c.font = FONT_HEADER if j > 0 else f()
    if j > 0:
        c.fill = FILL_HEADER
r += 1
FW_FIRST_DATA_ROW = r
# Platzhalter-Vergleichszeilen, die tatsächliche Objekt-Verknüpfung erfolgt aus 11_Kalkulation heraus
# (03_Foerderprogramme bleibt reine Konditions-Stammdatenbank; die Wegrechnung selbst
#  steht direkt auf 11_Kalkulation, weil sie objektspezifische Summen aus 10_VORLAGE braucht).
ws.cell(row=r, column=2, value="Hinweis: Die eigentliche Förderweg-Berechnung für ein konkretes Objekt steht auf dem Blatt 11_<ID> (Abschnitt 'Förderung'), da sie objektspezifische Sanierungssummen benötigt. Hier stehen nur die Konditionen.").font = FONT_COMMENT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

autofit(ws, {"A": 2, "B": 26, "C": 12, "D": 30, "E": 46, "F": 24, "G": 46})
print("03_Foerderprogramme OK")

# =====================================================================================
# 04_Nutzungsdauern
# =====================================================================================
ws = wb.create_sheet("04_Nutzungsdauern")
set_title(ws, "04_Nutzungsdauern — Zustandsfaktoren und Referenztabelle")
section(ws, 3, "Zustandsfaktor-Tabelle (verwendet in Checkliste 10_VORLAGE)", span=3)
header_row(ws, 4, ["Zustandsnote", "Zustandsfaktor", "Bedeutung"], col=2)
ZUSTAND = [(1, 1.10, "neuwertig"), (2, 1.00, "gut, altersgemäß"), (3, 0.70, "mittel, Gebrauchsspuren"),
           (4, 0.35, "schlecht, baldiger Handlungsbedarf"), (5, 0.00, "Sanierungsfall, sofort")]
ZUSTAND_FIRST_ROW = 5
for i, (note, faktor, bed) in enumerate(ZUSTAND):
    row = ZUSTAND_FIRST_ROW + i
    ws.cell(row=row, column=2, value=note).font = FONT_FORMULA
    c = ws.cell(row=row, column=3, value=faktor); c.font = FONT_FORMULA; c.number_format = "0.00"
    ws.cell(row=row, column=4, value=bed).font = FONT_FORMULA
ZUSTAND_LAST_ROW = ZUSTAND_FIRST_ROW + len(ZUSTAND) - 1

r = ZUSTAND_LAST_ROW + 2
section(ws, r, "Referenztabelle Nutzungsdauern (Dokumentation, BTE-Arbeitsblatt 2008 / VDI 2067)", span=3); r += 1
header_row(ws, r, ["Bauteil", "Jahre", "Quelle"], col=2); r += 1
REF_ND = [
    ("Dacheindeckung Ziegel/Beton", "48-60", "BTE-Arbeitsblatt 2008"),
    ("Flachdachabdichtung", "20", "BTE-Arbeitsblatt 2008"),
    ("Fenster Kunststoff", "37", "BTE-Arbeitsblatt 2008"),
    ("WDVS", "35 (konservativ) bis >50 (BBSR-Neufassung 2025)", "BTE 2008 / BMWSB-BBSR 2025"),
    ("Heizkessel Gas/Öl", "16-19", "BTE-Arbeitsblatt 2008"),
    ("Wärmepumpe", "15-20", "VDI 2067 / Praxis"),
    ("Elektroinstallation", "40-80", "ib-rauch.de"),
    ("Bad/Sanitärobjekte", "28-36", "BTE-Arbeitsblatt 2008"),
    ("Bodenbelag Parkett / PVC / Teppich", "44-61 / 21-28 / 6-12", "BTE-Arbeitsblatt 2008"),
]
for bauteil, jahre, quelle in REF_ND:
    ws.cell(row=r, column=2, value=bauteil).font = FONT_FORMULA
    ws.cell(row=r, column=3, value=jahre).font = FONT_FORMULA
    ws.cell(row=r, column=4, value=quelle).font = FONT_COMMENT
    r += 1

autofit(ws, {"A": 2, "B": 42, "C": 30, "D": 30})
print("04_Nutzungsdauern OK")

# =====================================================================================
# 05_Energiekennwerte
# =====================================================================================
ws = wb.create_sheet("05_Energiekennwerte")
set_title(ws, "05_Energiekennwerte — Baualtersklassen, PV-Staffelpreise")
section(ws, 3, "Baualtersklassen kWh/m²a (unsaniert) — genutzt wenn kein Energieausweis vorliegt", span=3)
header_row(ws, 4, ["Baujahr von", "Baujahr bis", "kWh/m²a"], col=2)
BAK = [(0, 1957, 275), (1958, 1978, 215), (1979, 1994, 160), (1995, 2009, 95), (2010, 2026, 60)]
BAK_FIRST_ROW = 5
for i, (von, bis, val) in enumerate(BAK):
    row = BAK_FIRST_ROW + i
    ws.cell(row=row, column=2, value=von).font = FONT_FORMULA
    ws.cell(row=row, column=3, value=bis).font = FONT_FORMULA
    ws.cell(row=row, column=4, value=val).font = FONT_FORMULA
BAK_LAST_ROW = BAK_FIRST_ROW + len(BAK) - 1

r = BAK_LAST_ROW + 2
section(ws, r, "Zielwerte nach Sanierung (Anhaltswerte, kein Nachweisersatz)", span=3); r += 1
header_row(ws, r, ["Zielstandard", "kWh/m²a"], col=2); r += 1
ZIEL_FIRST_ROW = r
ZIELWERTE = [("EH 55 EE", 65), ("EH 40 EE", 40), ("Nur Einzelmaßnahmen", 90)]
for name, val in ZIELWERTE:
    ws.cell(row=r, column=2, value=name).font = FONT_FORMULA
    ws.cell(row=r, column=3, value=val).font = FONT_FORMULA
    r += 1
ZIEL_LAST_ROW = r - 1

r += 1
section(ws, r, "PV-Staffelpreise (nicht linear)", span=3); r += 1
header_row(ws, r, ["kWp ab", "€/kWp"], col=2); r += 1
PV_STAFFEL_FIRST_ROW = r
PV_STAFFEL = [(0, 1500), (5, 1335), (10, 1200), (15, 1050), (20, 950)]
for kwp, preis in PV_STAFFEL:
    ws.cell(row=r, column=2, value=kwp).font = FONT_FORMULA
    ws.cell(row=r, column=3, value=preis).font = FONT_FORMULA
    r += 1
PV_STAFFEL_LAST_ROW = r - 1

autofit(ws, {"A": 2, "B": 16, "C": 16, "D": 16})
print("05_Energiekennwerte OK")

# =====================================================================================
# 10_VORLAGE — Objektstammdaten + Checkliste
# =====================================================================================
ws = wb.create_sheet("10_VORLAGE")
SHEET10 = "10_VORLAGE"
set_title(ws, "10_VORLAGE — Objektdaten und Sanierungs-Checkliste  [Beim Kopieren: Ziffernpräfix behalten, Rest durch Objekt-ID ersetzen]")

def local_name(ws, name, row, col=3):
    ref = f"'{ws.title}'!${get_column_letter(col)}${row}"
    ws.defined_names[name] = DefinedName(name, attr_text=ref)

r = 3
section(ws, r, "1. Objektstammdaten", span=4); r += 1
OBJ = {}  # semantic name -> row number (Wert-Spalte = C)

def obj_field(key, label, value, unit="", comment="", dropdown=None, fmt=None, is_input=True):
    global r
    label_value(ws, r, label, value, unit=unit, comment=comment, fmt=fmt, is_input=is_input,
                col_label=2, col_value=3, col_unit=4, col_comment=5)
    OBJ[key] = r
    local_name(ws, "Obj_" + key, r, col=3)
    if dropdown:
        add_dropdown(ws, f"C{r}", dropdown)
    r += 1

obj_field("Adresse", "Adresse", "Musterstraße 1")
obj_field("PLZ", "PLZ", 32312)
obj_field("Ort", "Ort", "Lübbecke")
obj_field("Baujahr", "Baujahr", 1978)
obj_field("Kaufpreis", "Kaufpreis (Angebot)", 285000, "€", fmt=EUR_FMT)
obj_field("Wohnflaeche", "Wohnfläche", 130, "m²")
obj_field("Grundstuecksflaeche", "Grundstücksfläche", 750, "m²")
obj_field("Bodenrichtwert", "Bodenrichtwert", 100, "€/m²", "BORIS-NRW / GMB 2026 prüfen", fmt=EUR2_FMT)
obj_field("Anpassungsfaktor", "Anpassungsfaktor Grundstücksgröße", 1.00, "Faktor", "GMB-Umrechnungskoeffizient, 400-1600m²-Gültigkeit")
obj_field("Geschosse", "Geschosse", 2)
obj_field("KellerTyp", "Keller", "Vollkeller", dropdown=["Kein Keller", "Teilkeller", "Vollkeller"])
obj_field("Dachform", "Dachform", "Satteldach", dropdown=["Satteldach", "Flachdach", "Walmdach"])
obj_field("DachEternit", "Dacheindeckung Eternit/Asbestverdacht", "Nein", dropdown=["Ja", "Nein"])
obj_field("Energieausweis", "Energieausweis-Wert (0 = kein Ausweis)", 185, "kWh/m²a")
obj_field("HeizungAlt", "Heizungsart aktuell", "Gasheizung",
          dropdown=["Gasheizung", "Ölheizung", "Nachtspeicher", "Fernwärme", "Wärmepumpe", "Sonstige"])
obj_field("AnzNachtspeicher", "Anzahl Nachtspeicheröfen", 0, "Stück")
obj_field("FeuchteKeller", "Feuchteschaden im Keller", "Nein", dropdown=["Ja", "Nein"])
obj_field("Zielstandard", "Zielstandard", "EH 55 EE", dropdown=["EH 55 EE", "EH 40 EE", "Nur Einzelmaßnahmen"])
obj_field("iSFP_JN", "iSFP-Bonus nutzen (Energieberater beauftragt)", "Ja", dropdown=["Ja", "Nein"],
           comment="steuert Deckel 30.000€/15% vs. 60.000€/20% bei BEG-EM")
obj_field("Foerderweg_Wahl", "Förderweg-Wahl", "Automatik (Empfehlung)",
           dropdown=["Automatik (Empfehlung)", "Weg A (Effizienzhaus)", "Weg B (Einzelmaßnahmen)"],
           comment="Empfehlung wird auf 11_Kalkulation berechnet, hier überschreibbar")
obj_field("MaklerJN", "Makler beteiligt", "Ja", dropdown=["Ja", "Nein"])
obj_field("AnzBaeder", "Anzahl Bäder", 2, "Stück")
obj_field("PersonenOverride", "Personen im Haushalt (0=Default aus Annahmen)", 0, "Anzahl")
obj_field("SanierungsdauerMonate", "Sanierungsdauer", 6, "Monate")
obj_field("EinzugNachSanierung", "Einzug erst nach Sanierung", "Ja", dropdown=["Ja", "Nein"])
obj_field("ParallelmieteMonat", "Miete während Sanierung", 900, "€/Monat", fmt=EUR_FMT)
obj_field("PV_kWp_Override", "PV Zielgröße (0 = automatisch aus Dachfläche)", 0, "kWp")
obj_field("PV_Ausrichtung", "PV Ausrichtung", "Süd", dropdown=["Süd", "Ost-West"])

# Grundstückswert (Formel)
r_gsw = r
label_value(ws, r, "Grundstückswert (berechnet)",
            f"=Obj_Grundstuecksflaeche*Obj_Bodenrichtwert*Obj_Anpassungsfaktor",
            unit="€", fmt=EUR_FMT, is_input=False)
ws.cell(row=r, column=3).font = FONT_RESULT
ws.cell(row=r, column=3).fill = FILL_RESULT
OBJ["Grundstueckswert"] = r
local_name(ws, "Obj_Grundstueckswert", r, col=3)
r += 2

section(ws, r, "2. Geometrie (automatisch berechnet, Zellen bei Bedarf mit Zahl überschreiben)", span=4); r += 1

def geo_field(key, label, formula, unit=""):
    global r
    label_value(ws, r, label, formula, unit=unit, is_input=False, col_label=2, col_value=3, col_unit=4)
    ws.cell(row=r, column=3).font = FONT_LINK
    OBJ[key] = r
    local_name(ws, "Obj_" + key, r, col=3)
    r += 1

geo_field("Grundflaeche", "Grundfläche", "=Obj_Wohnflaeche/Obj_Geschosse", "m²")
geo_field("Umfang", "Umfang", "=A_F_Umfang*SQRT(Obj_Grundflaeche)", "lfm")
geo_field("Fassadenflaeche", "Fassadenfläche", "=Obj_Umfang*Obj_Geschosse*A_F_Geschoss", "m²")
geo_field("Fensterflaeche", "Fensterfläche", "=Obj_Fassadenflaeche*A_F_Fenster", "m²")
geo_field("AnzFenster", "Anzahl Fenster", "=Obj_Fensterflaeche/A_F_Fenstergr", "Stück")
geo_field("Dachflaeche", "Dachfläche",
          '=IF(Obj_Dachform="Flachdach",Obj_Grundflaeche*A_F_Dach_flach,Obj_Grundflaeche*A_F_Dach)', "m²")
geo_field("AnzHeizkoerper", "Anzahl Heizkörper", "=ROUND(Obj_Wohnflaeche*A_F_HK,0)", "Stück")
geo_field("AnzTueren", "Anzahl Innentüren", "=ROUND(Obj_Wohnflaeche*A_F_Tueren,0)", "Stück")
geo_field("AnzRaeume", "Anzahl Räume", "=ROUND(Obj_Wohnflaeche*A_F_Raeume,0)", "Stück")
geo_field("PV_DachflaecheNutzbar", "PV nutzbare Dachfläche", "=Obj_Dachflaeche*A_PV_Anteil", "m²")
geo_field("PV_kWp", "PV-Anlagengröße",
          "=IF(Obj_PV_kWp_Override>0,Obj_PV_kWp_Override,Obj_PV_DachflaecheNutzbar/A_PV_Flaeche)", "kWp")
geo_field("PV_SpeicherKWh", "PV-Speichergröße", "=Obj_PV_kWp*A_Speicher_kWp", "kWh")
geo_field("PersonenEffektiv", "Personen im Haushalt (effektiv)",
          "=IF(Obj_PersonenOverride>0,Obj_PersonenOverride,A_Personen)", "Anzahl")

r += 1
print("10_VORLAGE Stammdaten/Geometrie OK, bis Zeile", r, "OBJ keys:", len(OBJ))

# ---------- 3. Sanierungs-Checkliste ----------
section(ws, r, "3. Sanierungs-Checkliste (eine Zeile je Maßnahme aus 02_Massnahmenkatalog)", span=20); r += 1

CHK_HEADERS = ["ID", "Kategorie", "Maßnahme", "Einheit", "Menge\nautomatisch", "Menge\nOverride",
               "Menge\neffektiv", "Einheits-\npreis €", "Kosten €", "Förder-\nfähig",
               "Baujahr/\nletzte Erneuerung", "Zustands-\nnote (1-5)", "Nutzungs-\ndauer (J)",
               "Zustands-\nfaktor", "Restlebens-\ndauer (J)", "Fälligkeits-\njahr",
               "Override\nZeitpunkt", "In Kalku-\nlation?", "Eigenleistung\n%", "Förderfähige\nKosten €",
               "Pflicht-Warnung", "Kommentar"]
CHK_HEADER_ROW = r
header_row(ws, r, CHK_HEADERS, col=2)
ws.row_dimensions[r].height = 45
r += 1
CHK_FIRST_ROW = r

COL = dict(ID=2, KAT=3, NAME=4, EINHEIT=5, MENGE_AUTO=6, MENGE_OVR=7, MENGE_EFF=8, PREIS=9,
           KOSTEN=10, FOERDERFAEHIG=11, BAUJAHR=12, ZUSTAND=13, NUTZUNG=14, ZFAKTOR=15,
           RESTLEBEN=16, FAELLIG=17, OVERRIDE_ZEIT=18, IN_KALK=19, EIGENLEISTUNG=20,
           FOERDERKOSTEN=21, PFLICHT_WARN=22, KOMMENTAR=23, PFLICHT55=24, PFLICHT40=25)

GENERIC_FK = {
    "dachflaeche": "Obj_Dachflaeche",
    "fassadenflaeche": "Obj_Fassadenflaeche",
    "fassadenflaeche_x2": "Obj_Fassadenflaeche*2",
    "fensterflaeche": "Obj_Fensterflaeche",
    "anz_fenster": "Obj_AnzFenster",
    "wohnflaeche": "Obj_Wohnflaeche",
    "wohnflaeche_fbh": "Obj_Wohnflaeche*A_F_FBH",
    "wohnflaeche_skaliert_wp": "Obj_Wohnflaeche/130",
    "wohnflaeche_skaliert_wp_erd": "Obj_Wohnflaeche/130",
    "grundflaeche": "Obj_Grundflaeche",
    "umfang": "Obj_Umfang",
    "anz_heizkoerper": "Obj_AnzHeizkoerper",
    "anz_tueren": "Obj_AnzTueren",
    "anz_raeume": "Obj_AnzRaeume",
    "anz_raeume_x1_5": "Obj_AnzRaeume*1.5",
    "anz_baeder": "Obj_AnzBaeder",
    "geschosse_minus1": "MAX(0,Obj_Geschosse-1)",
    "pauschal1": "1",
    "pauschal0": "0",
    "pv_kwp": "Obj_PV_kWp",
    "pv_speicher_kwh": "Obj_PV_SpeicherKWh",
    "pauschal1_prozent3": "1",
}

CONDITIONAL_OVERRIDES = {
    "H01": 'IF(Obj_Dachform="Satteldach",Obj_Dachflaeche,0)',
    "H02": 'IF(Obj_Dachform<>"Satteldach",Obj_Dachflaeche,0)',
    "H09": 'IF(Obj_KellerTyp<>"Kein Keller",Obj_Grundflaeche,0)',
    "H10": 'IF(Obj_KellerTyp="Vollkeller",Obj_Umfang,0)',
    "R01": 'IF(AND(Obj_Baujahr<1994,Obj_DachEternit="Ja"),Obj_Dachflaeche,0)',
    "R02": "IF(Obj_Baujahr<1994,Obj_Wohnflaeche,0)",
    "R03": "Obj_AnzNachtspeicher",
    "R04": 'IF(Obj_HeizungAlt="Ölheizung",1,0)',
    "R05": 'IF(Obj_FeuchteKeller="Ja",Obj_Umfang,0)',
    "R06": "IF(Obj_Baujahr<1994,1,0)",
    "A02": "0",
}

MK = "'02_Massnahmenkatalog'"
ND = "'04_Nutzungsdauern'"

for i, item in enumerate(CATALOG):
    row = CHK_FIRST_ROW + i
    cid = item["id"]
    B = f"$B{row}"
    ws.cell(row=row, column=COL["ID"], value=cid).font = FONT_FORMULA

    def lookup(col_letter):
        return f"INDEX({MK}!${col_letter}:${col_letter},MATCH({B},{MK}!$B:$B,0))"

    ws.cell(row=row, column=COL["KAT"], value=f"={lookup('C')}").font = FONT_LINK
    c = ws.cell(row=row, column=COL["NAME"], value=f"={lookup('D')}")
    c.font = FONT_LINK; c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row, column=COL["EINHEIT"], value=f"={lookup('E')}").font = FONT_LINK

    menge_expr = CONDITIONAL_OVERRIDES.get(cid, GENERIC_FK[item["fk"]])
    ws.cell(row=row, column=COL["MENGE_AUTO"], value=f"={menge_expr}").font = FONT_FORMULA

    ovr_cell = ws.cell(row=row, column=COL["MENGE_OVR"])
    ovr_cell.font = FONT_INPUT; ovr_cell.fill = FILL_INPUT

    auto_ref = f"${get_column_letter(COL['MENGE_AUTO'])}{row}"
    ovr_ref = f"${get_column_letter(COL['MENGE_OVR'])}{row}"
    ws.cell(row=row, column=COL["MENGE_EFF"],
            value=f'=IF({ovr_ref}="",{auto_ref},{ovr_ref})').font = FONT_FORMULA

    preis_cell_ref = f"${get_column_letter(COL['PREIS'])}{row}"
    if cid == "K01":
        ws.cell(row=row, column=COL["PREIS"]).value = (
            f"=LOOKUP($H{row},'05_Energiekennwerte'!$B${PV_STAFFEL_FIRST_ROW}:$B${PV_STAFFEL_LAST_ROW},"
            f"'05_Energiekennwerte'!$C${PV_STAFFEL_FIRST_ROW}:$C${PV_STAFFEL_LAST_ROW})")
    elif cid == "W02":
        first_j = f"${get_column_letter(COL['KOSTEN'])}${CHK_FIRST_ROW}"
        last_j = f"${get_column_letter(COL['KOSTEN'])}${CHK_FIRST_ROW + len(CATALOG) - 1}"
        first_c = f"${get_column_letter(COL['KAT'])}${CHK_FIRST_ROW}"
        last_c = f"${get_column_letter(COL['KAT'])}${CHK_FIRST_ROW + len(CATALOG) - 1}"
        ws.cell(row=row, column=COL["PREIS"]).value = (
            f'=0.03*(SUMIFS({first_j}:{last_j},{first_c}:{last_c},"Hülle")'
            f'+SUMIFS({first_j}:{last_j},{first_c}:{last_c},"Anlagentechnik"))')
    else:
        ws.cell(row=row, column=COL["PREIS"], value=f"={lookup('F')}")
    ws.cell(row=row, column=COL["PREIS"]).font = FONT_LINK
    ws.cell(row=row, column=COL["PREIS"]).number_format = EUR2_FMT

    eff_ref = f"${get_column_letter(COL['MENGE_EFF'])}{row}"
    ws.cell(row=row, column=COL["KOSTEN"], value=f"={eff_ref}*{preis_cell_ref}").font = FONT_FORMULA
    ws.cell(row=row, column=COL["KOSTEN"]).number_format = EUR_FMT

    ws.cell(row=row, column=COL["FOERDERFAEHIG"], value=f"={lookup('H')}").font = FONT_LINK
    ws.cell(row=row, column=COL["PFLICHT55"], value=f"={lookup('J')}").font = FONT_LINK
    ws.cell(row=row, column=COL["PFLICHT40"], value=f"={lookup('K')}").font = FONT_LINK

    bj_cell = ws.cell(row=row, column=COL["BAUJAHR"])
    bj_cell.font = FONT_INPUT; bj_cell.fill = FILL_INPUT
    zn_cell = ws.cell(row=row, column=COL["ZUSTAND"], value=3)
    zn_cell.font = FONT_INPUT; zn_cell.fill = FILL_INPUT
    add_dropdown(ws, f"{get_column_letter(COL['ZUSTAND'])}{row}", ["1", "2", "3", "4", "5"])

    ws.cell(row=row, column=COL["NUTZUNG"], value=f"={lookup('G')}").font = FONT_LINK

    zn_ref = f"${get_column_letter(COL['ZUSTAND'])}{row}"
    ws.cell(row=row, column=COL["ZFAKTOR"],
            value=f"=INDEX({ND}!$C:$C,MATCH({zn_ref},{ND}!$B:$B,0))").font = FONT_LINK

    nutz_ref = f"${get_column_letter(COL['NUTZUNG'])}{row}"
    zf_ref = f"${get_column_letter(COL['ZFAKTOR'])}{row}"
    bj_ref = f"${get_column_letter(COL['BAUJAHR'])}{row}"
    ws.cell(row=row, column=COL["RESTLEBEN"],
            value=f'={nutz_ref}*{zf_ref}-(A_Jahr-IF({bj_ref}="",Obj_Baujahr,{bj_ref}))').font = FONT_FORMULA

    rl_ref = f"${get_column_letter(COL['RESTLEBEN'])}{row}"
    ws.cell(row=row, column=COL["FAELLIG"], value=f"=A_Jahr+MAX(0,{rl_ref})").font = FONT_FORMULA

    ovz_cell = ws.cell(row=row, column=COL["OVERRIDE_ZEIT"], value="Automatik")
    ovz_cell.font = FONT_INPUT; ovz_cell.fill = FILL_INPUT
    add_dropdown(ws, f"{get_column_letter(COL['OVERRIDE_ZEIT'])}{row}",
                 ["Automatik", "Sofort", "In Horizont", "Später", "Entfällt"])

    ovz_ref = f"${get_column_letter(COL['OVERRIDE_ZEIT'])}{row}"
    faellig_ref = f"${get_column_letter(COL['FAELLIG'])}{row}"
    ws.cell(row=row, column=COL["IN_KALK"], value=(
        f'=IF({ovz_ref}="Automatik",IF({faellig_ref}<=A_Jahr+A_Horizont,"JA","NEIN"),'
        f'IF(OR({ovz_ref}="Sofort",{ovz_ref}="In Horizont"),"JA","NEIN"))')).font = FONT_FORMULA

    el_cell = ws.cell(row=row, column=COL["EIGENLEISTUNG"], value=0)
    el_cell.font = FONT_INPUT; el_cell.fill = FILL_INPUT; el_cell.number_format = PCT_FMT
    if not item["el"]:
        el_cell.value = 0

    kosten_ref = f"${get_column_letter(COL['KOSTEN'])}{row}"
    ff_ref = f"${get_column_letter(COL['FOERDERFAEHIG'])}{row}"
    el_ref = f"${get_column_letter(COL['EIGENLEISTUNG'])}{row}"
    inkalk_ref_u = f"${get_column_letter(COL['IN_KALK'])}{row}"
    ws.cell(row=row, column=COL["FOERDERKOSTEN"],
            value=f'=IF(AND({ff_ref}="Ja",{inkalk_ref_u}="JA"),{kosten_ref}*(1-{el_ref}),0)').font = FONT_FORMULA
    ws.cell(row=row, column=COL["FOERDERKOSTEN"]).number_format = EUR_FMT

    inkalk_ref = f"${get_column_letter(COL['IN_KALK'])}{row}"
    p55_ref = f"${get_column_letter(COL['PFLICHT55'])}{row}"
    p40_ref = f"${get_column_letter(COL['PFLICHT40'])}{row}"
    ws.cell(row=row, column=COL["PFLICHT_WARN"], value=(
        f'=IF(AND({inkalk_ref}="NEIN",OR(AND(Obj_Zielstandard="EH 55 EE",{p55_ref}="Ja"),'
        f'AND(Obj_Zielstandard="EH 40 EE",{p40_ref}="Ja"))),'
        f'"Pflicht bei Zielstandard, aber nicht aufgenommen","")')).font = FONT_WARN

    kom_cell = ws.cell(row=row, column=COL["KOMMENTAR"])
    kom_cell.font = FONT_INPUT; kom_cell.fill = FILL_INPUT

    for col_idx in range(2, COL["PFLICHT40"] + 1):
        ws.cell(row=row, column=col_idx).border = BORDER_ALL

CHK_LAST_ROW = CHK_FIRST_ROW + len(CATALOG) - 1
CHK_ROW_BY_ID = {item["id"]: CHK_FIRST_ROW + i for i, item in enumerate(CATALOG)}

# Bedingte Formatierung
rng_inkalk = f"{get_column_letter(COL['IN_KALK'])}{CHK_FIRST_ROW}:{get_column_letter(COL['IN_KALK'])}{CHK_LAST_ROW}"
ws.conditional_formatting.add(rng_inkalk, CellIsRule(operator="equal", formula=['"JA"'], fill=FILL_RESULT))
rng_warn = f"{get_column_letter(COL['PFLICHT_WARN'])}{CHK_FIRST_ROW}:{get_column_letter(COL['PFLICHT_WARN'])}{CHK_LAST_ROW}"
ws.conditional_formatting.add(rng_warn, FormulaRule(
    formula=[f'{get_column_letter(COL["PFLICHT_WARN"])}{CHK_FIRST_ROW}<>""'], fill=FILL_WARN))

# Hilfsspalten (Pflicht-Lookups) ausblenden
ws.column_dimensions[get_column_letter(COL["PFLICHT55"])].hidden = True
ws.column_dimensions[get_column_letter(COL["PFLICHT40"])].hidden = True

ws.auto_filter.ref = f"B{CHK_HEADER_ROW}:W{CHK_LAST_ROW}"

# WICHTIG: hier NICHT ws.freeze_panes = f"C{CHK_FIRST_ROW}" setzen. Freeze Panes
# fixiert IMMER alles ab Zeile 1 (nicht nur die Checkliste) - bei CHK_FIRST_ROW
# ~51 würde das den kompletten Stammdaten-Block als "eingefrorenen" Bereich
# behandeln. Da dieser Bereich mehr Zeilen hat, als in ein normales Fenster
# passen, bleibt für die eigentliche (scrollbare) Checkliste darunter kein
# sichtbarer Platz mehr übrig - sie wirkt dann komplett unsichtbar/nicht
# erreichbar, obwohl sie da ist (per Bugreport bestätigt). Stattdessen weiter
# unten (nach dem gesamten 10_VORLAGE-Aufbau) nur Zeile 1 + Spalte A/B fixieren,
# das lässt normales Scrollen über die komplette Fläche zu.

# ---------- Summenblock (Basis für 11_Kalkulation, Dashboard, Pie-Chart) ----------
r2 = CHK_LAST_ROW + 2
section(ws, r2, "4. Summen (für 11_Kalkulation)", span=4); r2 += 1
SUM_LABEL_ROW = r2
header_row(ws, r2, ["Kategorie", "Summe Kosten (nur In Kalkulation=JA) €", "davon förderfähig €"], col=2)
r2 += 1
SUM_FIRST_ROW = r2
KATEGORIEN = ["Hülle", "Anlagentechnik", "Elektro/Innenausbau", "Risiko/Schadstoff",
              "Klimaneutralität", "Smart Home", "Außenanlagen", "Weiche Kosten"]
kosten_range = f"${get_column_letter(COL['KOSTEN'])}${CHK_FIRST_ROW}:${get_column_letter(COL['KOSTEN'])}${CHK_LAST_ROW}"
foerder_range = f"${get_column_letter(COL['FOERDERKOSTEN'])}${CHK_FIRST_ROW}:${get_column_letter(COL['FOERDERKOSTEN'])}${CHK_LAST_ROW}"
kat_range = f"${get_column_letter(COL['KAT'])}${CHK_FIRST_ROW}:${get_column_letter(COL['KAT'])}${CHK_LAST_ROW}"
inkalk_range = f"${get_column_letter(COL['IN_KALK'])}${CHK_FIRST_ROW}:${get_column_letter(COL['IN_KALK'])}${CHK_LAST_ROW}"
for kat in KATEGORIEN:
    ws.cell(row=r2, column=2, value=kat).font = FONT_FORMULA
    ws.cell(row=r2, column=3, value=f'=SUMIFS({kosten_range},{kat_range},"{kat}",{inkalk_range},"JA")').font = FONT_FORMULA
    ws.cell(row=r2, column=3).number_format = EUR_FMT
    ws.cell(row=r2, column=4, value=f'=SUMIFS({foerder_range},{kat_range},"{kat}",{inkalk_range},"JA")').font = FONT_FORMULA
    ws.cell(row=r2, column=4).number_format = EUR_FMT
    r2 += 1
SUM_LAST_ROW = r2 - 1
CATROW = {kat: SUM_FIRST_ROW + i for i, kat in enumerate(KATEGORIEN)}
ws.cell(row=r2, column=2, value="GESAMT").font = FONT_RESULT
ws.cell(row=r2, column=3, value=f"=SUM(C{SUM_FIRST_ROW}:C{SUM_LAST_ROW})").font = FONT_RESULT
ws.cell(row=r2, column=3).fill = FILL_RESULT; ws.cell(row=r2, column=3).number_format = EUR_FMT
ws.cell(row=r2, column=4, value=f"=SUM(D{SUM_FIRST_ROW}:D{SUM_LAST_ROW})").font = FONT_RESULT
ws.cell(row=r2, column=4).fill = FILL_RESULT; ws.cell(row=r2, column=4).number_format = EUR_FMT
SUM_TOTAL_ROW = r2
# Eigenleistungs-Warnung
r2 += 2
el_range = f"${get_column_letter(COL['KOSTEN'])}${CHK_FIRST_ROW}:${get_column_letter(COL['KOSTEN'])}${CHK_LAST_ROW}"
el_pct_range = f"${get_column_letter(COL['EIGENLEISTUNG'])}${CHK_FIRST_ROW}:${get_column_letter(COL['EIGENLEISTUNG'])}${CHK_LAST_ROW}"
ws.cell(row=r2, column=2, value="Eigenleistungsanteil gesamt (Warnschwelle A_EL_Warn)").font = f()
ws.cell(row=r2, column=3, value=f"=SUMPRODUCT({el_range},{el_pct_range})/MAX(C{SUM_TOTAL_ROW},1)").font = FONT_FORMULA
ws.cell(row=r2, column=3).number_format = PCT_FMT
OBJ["EigenleistungGesamtPct"] = r2
EIGENLEISTUNG_WARN_ROW = r2

autofit(ws, {"A": 2, "B": 30, "C": 30, "D": 12, "E": 30, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12,
             "K": 10, "L": 14, "M": 10, "N": 10, "O": 10, "P": 12, "Q": 12, "R": 12, "S": 10,
             "T": 12, "U": 14, "V": 26, "W": 26})

# Nur Zeile 1 (Blatttitel) + Spalte A/B (ID/Kategorie-Spalten der Checkliste) fixieren -
# siehe Kommentar weiter oben, warum hier NICHT bis CHK_FIRST_ROW eingefroren wird.
ws.freeze_panes = "C2"

print(f"10_VORLAGE Checkliste OK: Zeilen {CHK_FIRST_ROW}-{CHK_LAST_ROW}, Summen {SUM_FIRST_ROW}-{SUM_TOTAL_ROW}")

# =====================================================================================
# Helper: Querverweise auf 10_VORLAGE
# =====================================================================================
def obj_ref(key):
    return f"'{SHEET10}'!$C${OBJ[key]}"

def chk_col_range(colkey):
    col = get_column_letter(COL[colkey])
    return f"'{SHEET10}'!${col}${CHK_FIRST_ROW}:${col}${CHK_LAST_ROW}"

def chk_cell(colkey, item_id):
    col = get_column_letter(COL[colkey])
    return f"'{SHEET10}'!${col}${CHK_ROW_BY_ID[item_id]}"

def sum_ref(kat, col_letter):
    return f"'{SHEET10}'!${col_letter}${CATROW[kat]}"

def heizung_sumifs(colkey):
    rng = chk_col_range(colkey)
    idrng = chk_col_range("ID")
    inkalkrng = chk_col_range("IN_KALK")
    return f'SUMIFS({rng},{idrng},"A01",{inkalkrng},"JA")+SUMIFS({rng},{idrng},"A02",{inkalkrng},"JA")'

# =====================================================================================
# 11_Kalkulation
# =====================================================================================
ws = wb.create_sheet("11_Kalkulation")
SHEET11 = "11_Kalkulation"
set_title(ws, "11_Kalkulation — Gesamtinvestition und Förderung  [Beim Kopieren: Ziffernpräfix behalten, Rest durch Objekt-ID ersetzen]")

r = 3
section(ws, r, "Kaufnebenkosten", span=4); r += 1
K = {}

def krow(label, formula, is_result=False, fmt=EUR_FMT, font_override=None):
    global r
    c_label = ws.cell(row=r, column=2, value=label); c_label.font = f(bold=is_result)
    c_val = ws.cell(row=r, column=3, value=formula)
    c_val.number_format = fmt
    c_val.font = font_override or (FONT_RESULT if is_result else FONT_FORMULA)
    if is_result:
        c_val.fill = FILL_RESULT
    K[label] = r
    r += 1
    return r - 1

ROW_KAUFPREIS = krow("Kaufpreis (Angebot)", f"={obj_ref('Kaufpreis')}", font_override=FONT_LINK)
ROW_GRESST = krow("+ Grunderwerbsteuer", f"=C{ROW_KAUFPREIS}*A_GrESt")
ROW_NOTAR = krow("+ Notar/Grundbuch", f"=C{ROW_KAUFPREIS}*A_Notar")
ROW_MAKLER = krow("+ Maklerprovision", f'=IF({obj_ref("MaklerJN")}="Ja",C{ROW_KAUFPREIS}*A_Makler,0)')
ROW_KP_NK = krow("→ Kaufpreis inkl. Nebenkosten", f"=SUM(C{ROW_KAUFPREIS}:C{ROW_MAKLER})", is_result=True)

r += 1
section(ws, r, "Sanierungskosten", span=4); r += 1
ROW_SAN_ROH = krow("Sanierungskosten Rohsumme (aus Checkliste)", f"='{SHEET10}'!$C${SUM_TOTAL_ROW}", font_override=FONT_LINK)
ROW_PUFFER = krow("+ Risikopuffer Unvorhergesehenes", f"=C{ROW_SAN_ROH}*A_Puffer")
ROW_SAN_GESAMT = krow("→ Sanierungskosten gesamt inkl. Puffer", f"=C{ROW_SAN_ROH}+C{ROW_PUFFER}", is_result=True)

r += 1
section(ws, r, "Förderweg-Vergleich (Konditionen siehe 03_Foerderprogramme)", span=4); r += 1
ROW_ZIELSTD = krow("Zielstandard (Anzeige)", f"={obj_ref('Zielstandard')}", fmt="General", font_override=FONT_LINK)

r += 1
ws.cell(row=r, column=2, value="Weg A — Effizienzhaus (KfW 261)").font = FONT_HEADER
ws.cell(row=r, column=2).fill = FILL_HEADER
r += 1
ROW_A_KREDIT = krow("Kreditbetrag Basis (Deckel 150.000€/WE)", f"=MIN(C{ROW_SAN_GESAMT},150000)")
ROW_A_SATZ = krow("Tilgungszuschuss-Satz",
                   f'=IF(C{ROW_ZIELSTD}="EH 40 EE",0.10,IF(C{ROW_ZIELSTD}="EH 55 EE",0.05,0))', fmt=PCT_FMT)
ROW_A_TILG = krow("Tilgungszuschuss", f"=C{ROW_A_KREDIT}*C{ROW_A_SATZ}")
ROW_A_ZINS = krow("Zinsvorteil (vereinfacht, nominal über Zinsbindung)",
                   f"=MAX(0,A_Zins_80-A_Zins_KfW261)*C{ROW_A_KREDIT}*A_Zinsbindung")
ROW_A_TOTAL = krow("→ Förderung Weg A gesamt", f"=C{ROW_A_TILG}+C{ROW_A_ZINS}", is_result=True)

r += 1
ws.cell(row=r, column=2, value="Weg B — Einzelmaßnahmen (BEG-EM + KfW 458)").font = FONT_HEADER
ws.cell(row=r, column=2).fill = FILL_HEADER
r += 1
huelle_topf_formula = (
    f"={sum_ref('Hülle','D')}+{sum_ref('Anlagentechnik','D')}+{sum_ref('Klimaneutralität','D')}"
    f"+{sum_ref('Smart Home','D')}-({heizung_sumifs('FOERDERKOSTEN')})"
)
ROW_B_HUELLE_TOPF = krow("Hülle-Topf förderfähige Kosten (ohne Heizung)", huelle_topf_formula, font_override=FONT_LINK)
ROW_B_DECKEL = krow("Deckel Hülle-Topf", f'=IF({obj_ref("iSFP_JN")}="Ja",60000,30000)')
ROW_B_SATZ = krow("Satz Hülle-Topf", f'=IF({obj_ref("iSFP_JN")}="Ja",0.20,0.15)', fmt=PCT_FMT)
ROW_B_ZUSCH_HUELLE = krow("Zuschuss Hülle-Topf", f"=MIN(C{ROW_B_HUELLE_TOPF},C{ROW_B_DECKEL})*C{ROW_B_SATZ}")
ROW_B_HEIZUNG_FOERD = krow("Heizungskosten förderfähig (Wärmepumpe A01/A02)",
                            f"={heizung_sumifs('FOERDERKOSTEN')}", font_override=FONT_LINK)
ROW_B_EINK_SATZ = krow("Einkommensbonus-Satz",
                        "=IF(A_Einkommen<=30000,0.40,IF(A_Einkommen<=40000,0.30,IF(A_Einkommen<=50000,0.10,0)))",
                        fmt=PCT_FMT)
ROW_B_SATZ_HEIZ = krow("Satz Heizung gesamt (gedeckelt)",
                        f"=MIN(0.30+0.16+C{ROW_B_EINK_SATZ},IF(A_Einkommen<=30000,0.80,0.70))", fmt=PCT_FMT)
ROW_B_ZUSCH_HEIZ = krow("Zuschuss Heizung (KfW 458, Deckel 28.000€)",
                         f"=MIN(C{ROW_B_HEIZUNG_FOERD},28000)*C{ROW_B_SATZ_HEIZ}")
ROW_B_ZUSCH_PLAN = krow("Zuschuss Fachplanung/Baubegleitung (50%, Deckel 5.000€)",
                         f"=MIN({chk_cell('KOSTEN','W02')}*0.5,5000)", font_override=FONT_LINK)
ROW_B_ZUSCH_BERAT = krow("Zuschuss Energieberatung/iSFP (50%, Deckel 650€)",
                          f"=MIN({chk_cell('KOSTEN','W01')}*0.5,650)", font_override=FONT_LINK)
ROW_B_ZINS = krow("Zinsvorteil KfW 358 (vereinfacht, nominal über Zinsbindung)",
                   f"=MAX(0,A_Zins_80-A_Zins_KfW358)*MIN(C{ROW_SAN_GESAMT},120000)*A_Zinsbindung")
ROW_B_TOTAL = krow("→ Förderung Weg B gesamt",
                    f"=C{ROW_B_ZUSCH_HUELLE}+C{ROW_B_ZUSCH_HEIZ}+C{ROW_B_ZUSCH_PLAN}+C{ROW_B_ZUSCH_BERAT}+C{ROW_B_ZINS}",
                    is_result=True)

r += 1
ROW_EMPFEHLUNG = krow("Empfehlung",
                       f'=IF(C{ROW_A_TOTAL}>C{ROW_B_TOTAL},"Weg A (Effizienzhaus)","Weg B (Einzelmaßnahmen)")',
                       fmt="General")
ROW_DIFFERENZ = krow("Differenz zwischen den Wegen", f"=ABS(C{ROW_A_TOTAL}-C{ROW_B_TOTAL})")
ROW_WAHL = krow("Gewählter Förderweg (Override in 10_VORLAGE)", f"={obj_ref('Foerderweg_Wahl')}",
                fmt="General", font_override=FONT_LINK)
wahl_ref = f"C{ROW_WAHL}"
ROW_FOERDERUNG_FINAL = krow(
    "→ Förderung gewählt (final)",
    f'=IF({wahl_ref}="Automatik (Empfehlung)",MAX(C{ROW_A_TOTAL},C{ROW_B_TOTAL}),'
    f'IF({wahl_ref}="Weg A (Effizienzhaus)",C{ROW_A_TOTAL},C{ROW_B_TOTAL}))', is_result=True)
ROW_KREDIT_BETRAG = krow(
    "Kreditbetrag Förderkredit (für 12_Finanzierung)",
    f'=IF(OR({wahl_ref}="Weg A (Effizienzhaus)",AND({wahl_ref}="Automatik (Empfehlung)",C{ROW_A_TOTAL}>=C{ROW_B_TOTAL})),'
    f"C{ROW_A_KREDIT},MIN(C{ROW_SAN_GESAMT},120000))")
ROW_KREDIT_ZINS = krow(
    "Zinssatz Förderkredit (für 12_Finanzierung)",
    f'=IF(OR({wahl_ref}="Weg A (Effizienzhaus)",AND({wahl_ref}="Automatik (Empfehlung)",C{ROW_A_TOTAL}>=C{ROW_B_TOTAL})),'
    "A_Zins_KfW261,A_Zins_KfW358)", fmt="0.00%")

r += 1
section(ws, r, "Gesamtinvestition", span=4); r += 1
ROW_NETTOINV = krow("Nettoinvestition (= Finanzierungsbedarf vor Eigenkapital)",
                     f"=C{ROW_KP_NK}+C{ROW_SAN_GESAMT}-C{ROW_FOERDERUNG_FINAL}", is_result=True)
kosten_rng = chk_col_range("KOSTEN")
faellig_rng = chk_col_range("FAELLIG")
inkalk_rng = chk_col_range("IN_KALK")
ROW_CASH_NOMINAL = krow(
    "Cash gesamt nominal indexiert (Lebenszyklus, alle In-Kalkulation-Maßnahmen)",
    f'=SUMPRODUCT(({inkalk_rng}="JA")*{kosten_rng}*(1+A_Baupreis)^({faellig_rng}-A_Jahr))')
ROW_BARWERT = krow(
    "Barwert Lebenszyklus (alle In-Kalkulation-Maßnahmen, auf heute abgezinst)",
    f'=SUMPRODUCT(({inkalk_rng}="JA")*{kosten_rng}*(1+A_Baupreis)^({faellig_rng}-A_Jahr)'
    f"/(1+A_Diskont)^({faellig_rng}-A_Jahr))")

r += 1
section(ws, r, "Datenbasis Kuchendiagramm", span=4); r += 1
PIE_FIRST = r
pie_items = [
    ("Grundstück", f"={obj_ref('Grundstueckswert')}"),
    ("Gebäude (Kaufpreis - Grundstück)", f"=C{ROW_KAUFPREIS}-{obj_ref('Grundstueckswert')}"),
    ("Kaufnebenkosten", f"=C{ROW_GRESST}+C{ROW_NOTAR}+C{ROW_MAKLER}"),
    ("Sanierung Hülle", f"={sum_ref('Hülle','C')}"),
    ("Sanierung Anlagentechnik", f"={sum_ref('Anlagentechnik','C')}"),
    ("Sanierung Elektro/Innenausbau", f"={sum_ref('Elektro/Innenausbau','C')}"),
    ("Klimaneutralität + Smart Home", f"={sum_ref('Klimaneutralität','C')}+{sum_ref('Smart Home','C')}"),
    ("Risiko/Schadstoff + Außenanlagen + Weiche Kosten",
     f"={sum_ref('Risiko/Schadstoff','C')}+{sum_ref('Außenanlagen','C')}+{sum_ref('Weiche Kosten','C')}"),
    ("Risikopuffer", f"=C{ROW_PUFFER}"),
]
for label, formula in pie_items:
    ws.cell(row=r, column=2, value=label).font = f()
    c = ws.cell(row=r, column=3, value=formula); c.font = FONT_FORMULA; c.number_format = EUR_FMT
    r += 1
PIE_LAST = r - 1

pie = PieChart()
pie.title = "Gesamtkosten nach Kategorie (vor Förderung)"
data = Reference(ws, min_col=3, min_row=PIE_FIRST, max_row=PIE_LAST)
cats = Reference(ws, min_col=2, min_row=PIE_FIRST, max_row=PIE_LAST)
pie.add_data(data, titles_from_data=False)
pie.set_categories(cats)
pie.height = 10
pie.width = 16
ws.add_chart(pie, f"F{PIE_FIRST}")

autofit(ws, {"A": 2, "B": 52, "C": 20})
print("11_Kalkulation OK, Nettoinvestition-Zeile:", ROW_NETTOINV)

# =====================================================================================
# 12_Finanzierung
# =====================================================================================
ws = wb.create_sheet("12_Finanzierung")
SHEET12 = "12_Finanzierung"
set_title(ws, "12_Finanzierung — Tranchen und Jahres-Tilgungsplan  [Beim Kopieren: Ziffernpräfix behalten, Rest durch Objekt-ID ersetzen]")
MAX_JAHRE = 35

r = 3
section(ws, r, "Tranchenlogik", span=4); r += 1
F = {}

def frow(label, formula, is_result=False, fmt=EUR_FMT, font_override=None):
    global r
    ws.cell(row=r, column=2, value=label).font = f(bold=is_result)
    c = ws.cell(row=r, column=3, value=formula)
    c.number_format = fmt
    c.font = font_override or (FONT_RESULT if is_result else FONT_FORMULA)
    if is_result:
        c.fill = FILL_RESULT
    F[label] = r
    r += 1
    return r - 1

ROW_NETTO_REF = frow("Nettoinvestition", f"='{SHEET11}'!$C${ROW_NETTOINV}", font_override=FONT_LINK)
ROW_EK = frow("- Eigenkapital", "=A_EK")
ROW_FIN_BEDARF = frow("→ Gesamtfinanzierungsbedarf", f"=MAX(0,C{ROW_NETTO_REF}-C{ROW_EK})", is_result=True)

r += 1
ws.cell(row=r, column=2, value="Tranche 1 — Förderkredit").font = FONT_HEADER
ws.cell(row=r, column=2).fill = FILL_HEADER
r += 1
ROW_T1_BETRAG = frow("Kreditbetrag Tranche 1",
                      f"=MIN(C{ROW_FIN_BEDARF},'{SHEET11}'!$C${ROW_KREDIT_BETRAG})")
ROW_T1_ZINS = frow("Zinssatz Tranche 1 (während Zinsbindung)",
                    f"='{SHEET11}'!$C${ROW_KREDIT_ZINS}", fmt="0.00%", font_override=FONT_LINK)
ROW_T1_TF = frow("Tilgungsfreie Jahre Tranche 1", "=A_KfW_tf", fmt="0")

r += 1
ws.cell(row=r, column=2, value="Tranche 2 — Hausbank-Annuitätendarlehen").font = FONT_HEADER
ws.cell(row=r, column=2).fill = FILL_HEADER
r += 1
ROW_T2_BETRAG = frow("Kreditbetrag Tranche 2 (Restbetrag)", f"=C{ROW_FIN_BEDARF}-C{ROW_T1_BETRAG}")
ROW_BLA = frow("Beleihungsauslauf (Näherung)",
               f"=C{ROW_FIN_BEDARF}/('{SHEET11}'!$C${ROW_KP_NK}+'{SHEET11}'!$C${ROW_SAN_GESAMT})", fmt=PCT_FMT)
ROW_T2_BASISZINS = frow("Basiszins nach Beleihungsauslauf",
                         f"=IF(C{ROW_BLA}<=0.7,A_Zins_60,IF(C{ROW_BLA}<=0.8,A_Zins_80,A_Zins_100))", fmt="0.00%")
ROW_T2_ZINS = frow(
    "Zinssatz Tranche 2 (abzgl. Grün-Rabatt bei EH-Zielstandard)",
    f"=C{ROW_T2_BASISZINS}-IF(OR('{SHEET11}'!$C${ROW_ZIELSTD}=\"EH 55 EE\",'{SHEET11}'!$C${ROW_ZIELSTD}=\"EH 40 EE\"),A_Zins_green,0)",
    fmt="0.00%")

r += 2
section(ws, r, "Kennzahlen-Kopf", span=4); r += 1
ROW_ANNUITAET1 = frow("Annuität Tranche 1 p.a. (nach Tilgungsfrei-Phase)", f"=C{ROW_T1_BETRAG}*(C{ROW_T1_ZINS}+A_Tilgung)")
ROW_ANNUITAET2 = frow("Annuität Tranche 2 p.a.", f"=C{ROW_T2_BETRAG}*(C{ROW_T2_ZINS}+A_Tilgung)")
ROW_RATE_MONAT = frow("Monatliche Gesamtrate (ab Ende Tilgungsfrei-Phase)",
                       f"=(C{ROW_ANNUITAET1}+C{ROW_ANNUITAET2})/12")

def build_tranche_schedule(start_row, title, betrag_ref, zins_ref, tf_ref, benchmark_ref, annuitaet_ref):
    """Baut einen 35-zeiligen Jahres-Tilgungsplan.
    Spalten: B=Jahr, C=Restschuld Start €, D=Zins %, E=Zins €, F=Tilgung €,
             G=Annuität €, H=Restschuld Ende €, I=Zinsvorteil vs. Marktzins €.
    Vereinfachung (dokumentiert): die Annuität (Zahlbetrag) bleibt über die
    gesamte Laufzeit konstant; nur der Zinssatz wechselt nach Ablauf der
    Zinsbindung auf A_Zins_Anschluss. Während der tilgungsfreien Jahre
    (tf_ref) entspricht die Zahlung exakt dem Zins (Annuität=Zins, Tilgung=0).
    """
    rr = start_row
    section(ws, rr, title, span=8); rr += 1
    header_row(ws, rr, ["Jahr", "Restschuld Start €", "Zins %", "Zins €", "Tilgung €",
                         "Annuität €", "Restschuld Ende €", "Zinsvorteil vs. Marktzins €"], col=2)
    header_rr = rr
    rr += 1
    first_rr = rr
    for i in range(1, MAX_JAHRE + 1):
        ws.cell(row=rr, column=2, value=i).number_format = "0"
        jahr_ref = f"$B{rr}"
        if i == 1:
            start_formula = f"={betrag_ref}"
        else:
            start_formula = f"=H{rr-1}"
        ws.cell(row=rr, column=3, value=start_formula).number_format = EUR_FMT
        start_ref = f"C{rr}"
        rate_formula = f"=IF({jahr_ref}<=A_Zinsbindung,{zins_ref},A_Zins_Anschluss)"
        ws.cell(row=rr, column=4, value=rate_formula).number_format = "0.00%"
        rate_ref = f"D{rr}"
        zins_eur_formula = f"=IF({jahr_ref}<=A_Laufzeit,{start_ref}*{rate_ref},0)"
        ws.cell(row=rr, column=5, value=zins_eur_formula).number_format = EUR_FMT
        zins_eur_ref = f"E{rr}"
        annuitaet_formula = f'=IF({jahr_ref}<=A_Laufzeit,IF({jahr_ref}<={tf_ref},{zins_eur_ref},{annuitaet_ref}),0)'
        ws.cell(row=rr, column=7, value=annuitaet_formula).number_format = EUR_FMT
        annuitaet_eur_ref = f"G{rr}"
        tilgung_formula = f"=IF({jahr_ref}<=A_Laufzeit,{annuitaet_eur_ref}-{zins_eur_ref},0)"
        ws.cell(row=rr, column=6, value=tilgung_formula).number_format = EUR_FMT
        tilgung_ref = f"F{rr}"
        restschuld_ende_formula = f"=MAX(0,{start_ref}-{tilgung_ref})"
        ws.cell(row=rr, column=8, value=restschuld_ende_formula).number_format = EUR_FMT
        vorteil_formula = f"=IF({jahr_ref}<=A_Laufzeit,{start_ref}*MAX(0,{benchmark_ref}-{rate_ref}),0)"
        ws.cell(row=rr, column=9, value=vorteil_formula).number_format = EUR_FMT
        for col in (2, 3, 4, 5, 6, 7, 8, 9):
            ws.cell(row=rr, column=col).font = FONT_FORMULA
        rr += 1
    return header_rr, first_rr, rr - 1

r += 1
T1_HDR, T1_FIRST, T1_LAST = build_tranche_schedule(
    r, "Jahres-Tilgungsplan Tranche 1 — Förderkredit (KfW 261)",
    betrag_ref=f"C{ROW_T1_BETRAG}", zins_ref=f"$C${ROW_T1_ZINS}", tf_ref="A_KfW_tf",
    benchmark_ref="$C$" + str(ROW_T2_BASISZINS), annuitaet_ref=f"$C${ROW_ANNUITAET1}")
r = T1_LAST + 2

T2_HDR, T2_FIRST, T2_LAST = build_tranche_schedule(
    r, "Jahres-Tilgungsplan Tranche 2 — Hausbank-Annuitätendarlehen",
    betrag_ref=f"C{ROW_T2_BETRAG}", zins_ref=f"$C${ROW_T2_ZINS}", tf_ref="0",
    benchmark_ref="$C$" + str(ROW_T2_ZINS), annuitaet_ref=f"$C${ROW_ANNUITAET2}")
r = T2_LAST + 2

section(ws, r, "Kennzahlen nach Tilgungsplan", span=4); r += 1
ROW_RESTSCHULD_ZB1 = frow(f"Restschuld Tranche 1 nach Zinsbindung (Jahr A_Zinsbindung)",
                           f"=INDEX(H{T1_FIRST}:H{T1_LAST},A_Zinsbindung)")
ROW_RESTSCHULD_ZB2 = frow(f"Restschuld Tranche 2 nach Zinsbindung (Jahr A_Zinsbindung)",
                           f"=INDEX(H{T2_FIRST}:H{T2_LAST},A_Zinsbindung)")
ROW_RESTSCHULD_ZB_GES = frow("→ Restschuld gesamt nach Zinsbindung (Anschlussfinanzierung nötig für)",
                              f"=C{ROW_RESTSCHULD_ZB1}+C{ROW_RESTSCHULD_ZB2}", is_result=True)
ROW_ZINSKOSTEN1 = frow("Gesamtzinskosten Tranche 1 über Laufzeit", f"=SUM(E{T1_FIRST}:E{T1_LAST})")
ROW_ZINSKOSTEN2 = frow("Gesamtzinskosten Tranche 2 über Laufzeit", f"=SUM(E{T2_FIRST}:E{T2_LAST})")
ROW_ZINSKOSTEN_GES = frow("→ Gesamtzinskosten Finanzierung über Laufzeit", f"=C{ROW_ZINSKOSTEN1}+C{ROW_ZINSKOSTEN2}",
                           is_result=True)
ROW_ZINSVORTEIL = frow("→ Zinsvorteil kumuliert durch Förderkredit ggü. marktüblicher Finanzierung",
                        f"=SUM(I{T1_FIRST}:I{T1_LAST})+SUM(I{T2_FIRST}:I{T2_LAST})", is_result=True)
ROW_MISCHZINS = frow("Effektiver Mischzins (anfänglich, gewichtet nach Kreditbetrag)",
                      f"=IF(C{ROW_FIN_BEDARF}=0,0,(C{ROW_T1_BETRAG}*C{ROW_T1_ZINS}+C{ROW_T2_BETRAG}*C{ROW_T2_ZINS})/C{ROW_FIN_BEDARF})",
                      fmt="0.00%")
ROW_RESTSCHULD_ENDLAUFZEIT = frow(
    "→ Restschuld gesamt am Ende der Gesamtlaufzeit (ggf. Anschlussfinanzierung/Sondertilgung nötig)",
    f"=INDEX(H{T1_FIRST}:H{T1_LAST},A_Laufzeit)+INDEX(H{T2_FIRST}:H{T2_LAST},A_Laufzeit)", is_result=True)

ws.cell(row=ROW_RESTSCHULD_ZB1, column=5,
        value="Hinweis: Annuität bleibt über die Laufzeit konstant; nur der Zins wechselt nach Zinsbindung "
              "auf A_Zins_Anschluss (konservative Annahme, siehe 01_Annahmen).").font = FONT_COMMENT
ws.cell(row=ROW_RESTSCHULD_ENDLAUFZEIT, column=5,
        value="Wichtig: Da die Annuität nach Zinsbindungsablauf NICHT neu berechnet wird (konservative "
              "Modellannahme), kann bei einem Zinsanstieg eine Restschuld am Laufzeitende verbleiben. "
              "In der Praxis würde die Bank die Rate bei Anschlussfinanzierung anpassen, um fristgerecht "
              "zu tilgen — ein Wert > 0 hier zeigt, wie groß dieser Anpassungsbedarf wäre.").font = FONT_WARN

autofit(ws, {"A": 2, "B": 8, "C": 16, "D": 10, "E": 14, "F": 14, "G": 14, "H": 16, "I": 20})
print("12_Finanzierung OK. Tranche1:", T1_FIRST, "-", T1_LAST, " Tranche2:", T2_FIRST, "-", T2_LAST)

wb.save(OUT)
print("Zwischenspeichern OK:", OUT)

# =====================================================================================
# 13_Bauzeit — Liquidität während der Sanierung
# =====================================================================================
ws = wb.create_sheet("13_Bauzeit")
SHEET13 = "13_Bauzeit"
set_title(ws, "13_Bauzeit — Liquidität während der Sanierung  [Beim Kopieren: Ziffernpräfix behalten, Rest durch Objekt-ID ersetzen]")

r = 3
section(ws, r, "Rahmendaten", span=4); r += 1

def brow(label, formula, is_result=False, fmt=EUR_FMT, font_override=None, is_input=False):
    global r
    ws.cell(row=r, column=2, value=label).font = f(bold=is_result)
    c = ws.cell(row=r, column=3, value=formula)
    c.number_format = fmt
    if is_input:
        c.font = FONT_INPUT; c.fill = FILL_INPUT
    else:
        c.font = font_override or (FONT_RESULT if is_result else FONT_FORMULA)
        if is_result:
            c.fill = FILL_RESULT
    r += 1
    return r - 1

ROW_BZ_DAUER = brow("Sanierungsdauer (Monate)", f"={obj_ref('SanierungsdauerMonate')}", fmt="0", font_override=FONT_LINK)
ROW_BZ_SANKOSTEN = brow("Sanierungskosten gesamt (Referenz)", f"='{SHEET11}'!$C${ROW_SAN_GESAMT}", font_override=FONT_LINK)
ROW_BZ_KREDIT = brow("Gesamtkreditbetrag (Referenz)", f"='{SHEET12}'!$C${ROW_FIN_BEDARF}", font_override=FONT_LINK)

r += 1
section(ws, r, "Monatlicher Zahlungs- und Abrufplan (12 Monate, editierbar)", span=7); r += 1
header_row(ws, r, ["Monat", "Anteil Sanierungssumme %", "Kosten Monat €", "Kumulierte Kosten €",
                    "Nicht abgerufener Kreditteil €", "Bereitstellungszins €", "Eigenleistung Std./Monat"], col=2)
r += 1
BZ_FIRST = r
DEFAULT_ANTEILE = [0.20, 0.15, 0.15, 0.15, 0.15, 0.20, 0, 0, 0, 0, 0, 0]
for i in range(12):
    row = r
    mc = ws.cell(row=row, column=2, value=i + 1)
    mc.number_format = "0"; mc.font = FONT_FORMULA
    anteil_cell = ws.cell(row=row, column=3, value=DEFAULT_ANTEILE[i])
    anteil_cell.font = FONT_INPUT; anteil_cell.fill = FILL_INPUT; anteil_cell.number_format = PCT_FMT
    kc = ws.cell(row=row, column=4, value=f"=C{row}*$C${ROW_BZ_SANKOSTEN}")
    kc.font = FONT_FORMULA; kc.number_format = EUR_FMT
    kum_formula = f"=D{row}" if i == 0 else f"=E{row-1}+D{row}"
    kumc = ws.cell(row=row, column=5, value=kum_formula)
    kumc.font = FONT_FORMULA; kumc.number_format = EUR_FMT
    na_formula = f"=$C${ROW_BZ_KREDIT}" if i == 0 else f"=MAX(0,$C${ROW_BZ_KREDIT}-E{row-1})"
    nac = ws.cell(row=row, column=6, value=na_formula)
    nac.font = FONT_FORMULA; nac.number_format = EUR_FMT
    bc = ws.cell(row=row, column=7, value=f"=IF(B{row}<=A_Bereit_frei,0,F{row}*A_Bereitstellung)")
    bc.font = FONT_FORMULA; bc.number_format = EUR_FMT
    el_cell = ws.cell(row=row, column=8, value=0)
    el_cell.font = FONT_INPUT; el_cell.fill = FILL_INPUT; el_cell.number_format = "0"
    r += 1
BZ_LAST = r - 1

ws.cell(row=r, column=2, value="Summe / Anteil-Check").font = FONT_RESULT
sum_anteil_cell = ws.cell(row=r, column=3, value=f"=SUM(C{BZ_FIRST}:C{BZ_LAST})")
sum_anteil_cell.font = FONT_RESULT; sum_anteil_cell.number_format = PCT_FMT
sum_kosten_cell = ws.cell(row=r, column=4, value=f"=SUM(D{BZ_FIRST}:D{BZ_LAST})")
sum_kosten_cell.font = FONT_RESULT; sum_kosten_cell.number_format = EUR_FMT
sum_bereit_cell = ws.cell(row=r, column=7, value=f"=SUM(G{BZ_FIRST}:G{BZ_LAST})")
sum_bereit_cell.font = FONT_RESULT; sum_bereit_cell.number_format = EUR_FMT
ROW_BZ_ANTEIL_CHECK = r
ROW_BZ_BEREIT_SUMME = r
r += 1
warn_cell = ws.cell(row=r, column=2,
    value=f'=IF(ABS(C{ROW_BZ_ANTEIL_CHECK}-1)>0.001,"WARNUNG: Anteile summieren nicht auf 100%","")')
warn_cell.font = FONT_WARN
r += 2

section(ws, r, "Zusatzkosten Bauzeit (informativ, nicht in 11_Kalkulation enthalten)", span=4); r += 1
ROW_BZ_PARALLELMIETE = brow("Parallelmiete gesamt",
    f'=IF({obj_ref("EinzugNachSanierung")}="Ja",{obj_ref("ParallelmieteMonat")}*{obj_ref("SanierungsdauerMonate")},0)')
ROW_BZ_BEREIT_GESAMT = brow("Bereitstellungszinsen gesamt", f"=G{ROW_BZ_BEREIT_SUMME}")
ROW_BZ_GESAMT = brow("→ Zusatzkosten Bauzeit gesamt (zusätzlicher Liquiditätsbedarf)",
    f"=C{ROW_BZ_PARALLELMIETE}+C{ROW_BZ_BEREIT_GESAMT}", is_result=True)
ws.cell(row=ROW_BZ_GESAMT, column=5,
    value="Hinweis: separat ausgewiesen, da abhängig vom Bauablauf; bei Bedarf manuell zum Eigenkapitalbedarf "
          "auf 12_Finanzierung addieren (bewusst nicht automatisch verrechnet, um eine Zirkelbeziehung mit "
          "der Kreditbetrags-Berechnung zu vermeiden).").font = FONT_COMMENT

r += 1
section(ws, r, "Eigenleistung-Terminrisiko (informativ, kein Cash-Effekt)", span=4); r += 1
ROW_BZ_EL_MAX = brow("Maximale Eigenleistungsstunden in einem Monat", f"=MAX(H{BZ_FIRST}:H{BZ_LAST})", fmt="0")
ROW_BZ_EL_WARN = brow("Terminrisiko-Hinweis",
    f'=IF(C{ROW_BZ_EL_MAX}>60,"Terminrisiko: über 60h Eigenleistung in einem Monat geplant - realistisch neben Beruf pruefen?","unauffällig")',
    fmt="General")

bar = BarChart()
bar.title = "Monatlicher Cash-out während der Sanierung"
bar.y_axis.title = "€"
bar.x_axis.title = "Monat"
bz_data = Reference(ws, min_col=4, min_row=BZ_FIRST - 1, max_row=BZ_LAST)
bz_cats = Reference(ws, min_col=2, min_row=BZ_FIRST, max_row=BZ_LAST)
bar.add_data(bz_data, titles_from_data=True)
bar.set_categories(bz_cats)
bar.height = 8
bar.width = 16
ws.add_chart(bar, f"J{BZ_FIRST}")

autofit(ws, {"A": 2, "B": 26, "C": 16, "D": 14, "E": 16, "F": 20, "G": 16, "H": 18})
print("13_Bauzeit OK")

# =====================================================================================
# 14_Betriebskosten — Vorher/Nachher
# =====================================================================================
ws = wb.create_sheet("14_Betriebskosten")
SHEET14 = "14_Betriebskosten"
set_title(ws, "14_Betriebskosten — Energiekosten vorher/nachher  [Beim Kopieren: Ziffernpräfix behalten, Rest durch Objekt-ID ersetzen]")

r = 3
section(ws, r, "Heizungsart-Preis- und CO2-Zuordnung (automatisch aus Objektstammdaten)", span=4); r += 1

OR_HEIZUNGALT = obj_ref("HeizungAlt")
OR_WOHNFLAECHE = obj_ref("Wohnflaeche")
OR_BAUJAHR = obj_ref("Baujahr")
OR_ENERGIEAUSWEIS = obj_ref("Energieausweis")
OR_ZIELSTANDARD = obj_ref("Zielstandard")
OR_PERSONENEFF = obj_ref("PersonenEffektiv")
OR_PVKWP = obj_ref("PV_kWp")
OR_PVAUSRICHTUNG = obj_ref("PV_Ausrichtung")

def heiz_price_formula():
    return (f'=IF({OR_HEIZUNGALT}="Gasheizung",A_P_Gas,'
            f'IF({OR_HEIZUNGALT}="Ölheizung",A_P_Oel,'
            f'IF({OR_HEIZUNGALT}="Nachtspeicher",A_P_Strom,'
            f'IF({OR_HEIZUNGALT}="Fernwärme",A_P_Fernwaerme,'
            f'IF({OR_HEIZUNGALT}="Wärmepumpe",A_P_WPStrom,A_P_Gas)))))')

def heiz_co2_formula():
    return (f'=IF({OR_HEIZUNGALT}="Gasheizung",A_CO2_Gas,'
            f'IF({OR_HEIZUNGALT}="Ölheizung",A_CO2_Oel,'
            f'IF({OR_HEIZUNGALT}="Nachtspeicher",A_CO2_Strom,'
            f'IF({OR_HEIZUNGALT}="Fernwärme",A_CO2_Fernwaerme,'
            f'IF({OR_HEIZUNGALT}="Wärmepumpe",A_CO2_Strom,A_CO2_Gas)))))')

def erow(label, formula, is_result=False, fmt=EUR_FMT, font_override=None, unit=None):
    global r
    ws.cell(row=r, column=2, value=label).font = f(bold=is_result)
    c = ws.cell(row=r, column=3, value=formula)
    c.number_format = fmt
    c.font = font_override or (FONT_RESULT if is_result else FONT_FORMULA)
    if is_result:
        c.fill = FILL_RESULT
    if unit:
        ws.cell(row=r, column=4, value=unit).font = FONT_COMMENT
    r += 1
    return r - 1

ROW_E_PREIS_ALT = erow("Energiepreis aktuelle Heizungsart", heiz_price_formula(), fmt=EUR2_FMT, unit="€/kWh")
ROW_E_CO2_ALT = erow("CO2-Faktor aktuelle Heizungsart", heiz_co2_formula(), fmt="0", unit="g/kWh")

r += 1
section(ws, r, "Vorher (unsaniert / Ist-Zustand)", span=4); r += 1
ROW_E_KENNWERT_ALT = erow("Energiekennwert (Ausweis oder Baualtersklasse)",
    f"=IF({OR_ENERGIEAUSWEIS}>0,{OR_ENERGIEAUSWEIS},LOOKUP({OR_BAUJAHR},"
    f"'05_Energiekennwerte'!$B${BAK_FIRST_ROW}:$B${BAK_LAST_ROW},"
    f"'05_Energiekennwerte'!$D${BAK_FIRST_ROW}:$D${BAK_LAST_ROW}))",
    fmt="0", unit="kWh/m²a")
ROW_E_VERBRAUCH_ALT = erow("Verbrauch vorher", f"=C{ROW_E_KENNWERT_ALT}*{OR_WOHNFLAECHE}", fmt="#,##0", unit="kWh/a")
ROW_E_KOSTEN_ALT = erow("Energiekosten vorher", f"=C{ROW_E_VERBRAUCH_ALT}*C{ROW_E_PREIS_ALT}", is_result=True)

r += 1
section(ws, r, "Nachher (nach Sanierung auf Zielstandard)", span=4); r += 1
ROW_E_ZIELWERT = erow("Zielwert Heizwärmebedarf",
    f"=INDEX('05_Energiekennwerte'!$C${ZIEL_FIRST_ROW}:$C${ZIEL_LAST_ROW},"
    f"MATCH({OR_ZIELSTANDARD},'05_Energiekennwerte'!$B${ZIEL_FIRST_ROW}:$B${ZIEL_LAST_ROW},0))",
    fmt="0", unit="kWh/m²a")
ROW_E_HEIZWAERME_NACHHER = erow("Heizwärmebedarf nachher", f"=C{ROW_E_ZIELWERT}*{OR_WOHNFLAECHE}", fmt="#,##0", unit="kWh/a")
ROW_E_WW = erow("Warmwasserbedarf", f"={OR_PERSONENEFF}*A_WW_Person", fmt="#,##0", unit="kWh/a")
ROW_E_WAERME_GESAMT = erow("Wärmebedarf gesamt (Heizung+WW)", f"=C{ROW_E_HEIZWAERME_NACHHER}+C{ROW_E_WW}",
    fmt="#,##0", unit="kWh/a")
ROW_E_JAZ = erow("Jahresarbeitszahl Wärmepumpe (gemäß gewählter Heizflächen)",
    f'=IF({chk_cell("IN_KALK", "A03")}="JA",A_JAZ_FBH,IF({chk_cell("IN_KALK", "A04")}="JA",A_JAZ_HK45,A_JAZ_HK55))',
    fmt="0.00", unit="-")
ROW_E_STROM_WP = erow("Strombedarf Wärmepumpe", f"=C{ROW_E_WAERME_GESAMT}/C{ROW_E_JAZ}", fmt="#,##0", unit="kWh/a")
ROW_E_STROM_HH = erow("Haushaltsstrombedarf (ohne Heizung)", f"={OR_PERSONENEFF}*A_HH_Strom_Person",
    fmt="#,##0", unit="kWh/a")
ROW_E_STROM_GESAMT = erow("Gesamtstrombedarf Haus", f"=C{ROW_E_STROM_WP}+C{ROW_E_STROM_HH}", fmt="#,##0", unit="kWh/a")

r += 1
section(ws, r, "Photovoltaik-Ertrag und -Nutzung", span=4); r += 1
ROW_E_PV_ERTRAG = erow("PV-Ertrag", f'={OR_PVKWP}*A_PV_Ertrag*IF({OR_PVAUSRICHTUNG}="Ost-West",A_PV_OW,1)',
    fmt="#,##0", unit="kWh/a")
ROW_E_PV_EIGEN = erow("PV-Eigenverbrauch", f"=MIN(C{ROW_E_PV_ERTRAG}*A_PV_EV,C{ROW_E_STROM_GESAMT})",
    fmt="#,##0", unit="kWh/a")
ROW_E_PV_UEBERSCHUSS = erow("PV-Überschuss (Einspeisung)", f"=C{ROW_E_PV_ERTRAG}-C{ROW_E_PV_EIGEN}",
    fmt="#,##0", unit="kWh/a")
ROW_E_NETZBEZUG = erow("Netzbezug", f"=C{ROW_E_STROM_GESAMT}-C{ROW_E_PV_EIGEN}", fmt="#,##0", unit="kWh/a")

r += 1
section(ws, r, "Ergebnis", span=4); r += 1
ROW_E_KOSTEN_NACHHER = erow("Energiekosten nachher",
    f"=C{ROW_E_NETZBEZUG}*A_P_WPStrom-C{ROW_E_PV_UEBERSCHUSS}*A_P_Einspeisung", is_result=True)
ROW_E_ERSPARNIS = erow("→ Jährliche Ersparnis", f"=C{ROW_E_KOSTEN_ALT}-C{ROW_E_KOSTEN_NACHHER}", is_result=True)

r += 1
section(ws, r, "20-Jahres-Verlauf (mit Energiepreissteigerung, Barwert)", span=5); r += 1
header_row(ws, r, ["Jahr", "Kosten vorher (nominal) €", "Kosten nachher (nominal) €", "Ersparnis nominal €",
                    "Ersparnis Barwert €"], col=2)
r += 1
EV_FIRST = r
for i in range(1, 21):
    row = r
    yc = ws.cell(row=row, column=2, value=i); yc.number_format = "0"; yc.font = FONT_FORMULA
    c3 = ws.cell(row=row, column=3, value=f"=$C${ROW_E_KOSTEN_ALT}*(1+A_Energiepreis)^(B{row}-1)")
    c3.number_format = EUR_FMT; c3.font = FONT_FORMULA
    c4 = ws.cell(row=row, column=4, value=f"=$C${ROW_E_KOSTEN_NACHHER}*(1+A_Energiepreis)^(B{row}-1)")
    c4.number_format = EUR_FMT; c4.font = FONT_FORMULA
    c5 = ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
    c5.number_format = EUR_FMT; c5.font = FONT_FORMULA
    c6 = ws.cell(row=row, column=6, value=f"=E{row}/(1+A_Diskont)^B{row}")
    c6.number_format = EUR_FMT; c6.font = FONT_FORMULA
    r += 1
EV_LAST = r - 1
ROW_E_ERSPARNIS_BARWERT = erow("→ Ersparnis über 20 Jahre, Barwert kumuliert", f"=SUM(F{EV_FIRST}:F{EV_LAST})",
    is_result=True)
ROW_E_ERSPARNIS_NOMINAL = erow("Ersparnis über 20 Jahre, nominal kumuliert", f"=SUM(E{EV_FIRST}:E{EV_LAST})")

bar2 = BarChart()
bar2.title = "Energiekosten vorher/nachher über 20 Jahre (nominal)"
bar2.y_axis.title = "€/a"
bar2.x_axis.title = "Jahr"
ev_data = Reference(ws, min_col=3, max_col=4, min_row=EV_FIRST - 1, max_row=EV_LAST)
ev_cats = Reference(ws, min_col=2, min_row=EV_FIRST, max_row=EV_LAST)
bar2.add_data(ev_data, titles_from_data=True)
bar2.set_categories(ev_cats)
bar2.height = 9
bar2.width = 18
ws.add_chart(bar2, f"H{EV_FIRST}")

autofit(ws, {"A": 2, "B": 42, "C": 18, "D": 18, "E": 16, "F": 16})
print("14_Betriebskosten OK")

# =====================================================================================
# 15_Zielpreis — Rückwärtsrechnung Kaufpreisobergrenze
# =====================================================================================
ws = wb.create_sheet("15_Zielpreis")
SHEET15 = "15_Zielpreis"
set_title(ws, "15_Zielpreis — Rückwärtsrechnung Kaufpreisobergrenze  [Beim Kopieren: Ziffernpräfix behalten, Rest durch Objekt-ID ersetzen]")

r = 3
section(ws, r, "Neubau-Referenzkosten (nur €/m² Wohnfläche)", span=4); r += 1

def zrow(label, formula, is_result=False, fmt=EUR_FMT, font_override=None):
    global r
    ws.cell(row=r, column=2, value=label).font = f(bold=is_result)
    c = ws.cell(row=r, column=3, value=formula)
    c.number_format = fmt
    c.font = font_override or (FONT_RESULT if is_result else FONT_FORMULA)
    if is_result:
        c.fill = FILL_RESULT
    r += 1
    return r - 1

OR15_ZIELSTANDARD = obj_ref("Zielstandard")
OR15_WOHNFLAECHE = obj_ref("Wohnflaeche")

ROW_Z_MEHRKOSTEN = zrow("Mehrkosten je Zielstandard",
    f'=IF({OR15_ZIELSTANDARD}="EH 40 EE",A_NB_Mehrkosten_EH40,IF({OR15_ZIELSTANDARD}="EH 55 EE",A_NB_Mehrkosten_EH55,0))',
    fmt=EUR2_FMT)
ROW_Z_NBGEB = zrow("Neubaukosten Gebäude (KG300+400)",
    f"={OR15_WOHNFLAECHE}*(A_NB_GEG+C{ROW_Z_MEHRKOSTEN})*A_Regio")
ROW_Z_NBGES = zrow("→ Neubaukosten gesamt (inkl. Baunebenkosten, Außenanlagen, Anschlüsse)",
    f"=C{ROW_Z_NBGEB}*(1+A_BNK)+{OR15_WOHNFLAECHE}*A_Aussen+A_Anschluesse", is_result=True)

r += 1
section(ws, r, "Zielwert gebrauchtes Gebäude (Neubau abzgl. Zielrabatt)", span=4); r += 1
ROW_Z_ZIELWERT = zrow("Zielwert Gebäude gebraucht (Neubau × (1 - Zielrabatt))",
    f"=C{ROW_Z_NBGES}*(1-A_Zielrabatt)", is_result=True)

r += 1
section(ws, r, "Rückrechnung auf Kaufpreisobergrenze", span=4); r += 1
ROW_Z_SANNETTO = zrow("Sanierungskosten gesamt, netto nach Förderung",
    f"='{SHEET11}'!$C${ROW_SAN_GESAMT}-'{SHEET11}'!$C${ROW_FOERDERUNG_FINAL}", font_override=FONT_LINK)
ROW_Z_MAXGEBAEUDE = zrow("Max. Kaufpreis Gebäudeanteil", f"=C{ROW_Z_ZIELWERT}-C{ROW_Z_SANNETTO}")
ROW_Z_GRUNDSTUECK = zrow("+ Grundstückswert (Bodenrichtwert × Fläche)", f"={obj_ref('Grundstueckswert')}",
    font_override=FONT_LINK)
ROW_Z_MAXGESAMT = zrow("→ Max. akzeptabler Kaufpreis gesamt", f"=C{ROW_Z_MAXGEBAEUDE}+C{ROW_Z_GRUNDSTUECK}",
    is_result=True)
ROW_Z_ANGEBOT = zrow("Angebotspreis (Referenz)", f"={obj_ref('Kaufpreis')}", font_override=FONT_LINK)
ROW_Z_DELTA = zrow("→ Verhandlungsdelta (positiv = Angebot ist günstig genug)",
    f"=C{ROW_Z_MAXGESAMT}-C{ROW_Z_ANGEBOT}", is_result=True)
ws.cell(row=ROW_Z_DELTA, column=5,
    value="Grundstück wird bewusst vor dem Prozentvergleich herausgerechnet und danach wieder addiert, da es "
          "bei Kauf und fiktivem Neubau gleich teuer ist (siehe Bauplan Abschnitt 11).").font = FONT_COMMENT

r += 2
section(ws, r, "Sensitivität: Max. Kaufpreis nach Zielrabatt × Regionalfaktor (fest berechnet, kein Excel-Datatable)",
        span=5); r += 1
ws.cell(row=r, column=2, value="Zielrabatt \\ Regionalfaktor").font = FONT_HEADER
ws.cell(row=r, column=2).fill = FILL_HEADER
REGIOFAKTOREN = [0.90, 1.00, 1.10]
for j, rf in enumerate(REGIOFAKTOREN):
    hc = ws.cell(row=r, column=3 + j, value=rf)
    hc.font = FONT_HEADER; hc.fill = FILL_HEADER; hc.number_format = "0.00"
r += 1
SENS_FIRST_ROW = r
ZIELRABATTE = [0.00, 0.10, 0.20, 0.30]
for i, zr_val in enumerate(ZIELRABATTE):
    row = r
    rc = ws.cell(row=row, column=2, value=zr_val)
    rc.font = FONT_FORMULA; rc.number_format = PCT_FMT
    for j, rf in enumerate(REGIOFAKTOREN):
        col = 3 + j
        formula = (f"=(({OR15_WOHNFLAECHE}*(A_NB_GEG+C{ROW_Z_MEHRKOSTEN})*{rf})*(1+A_BNK)"
                   f"+{OR15_WOHNFLAECHE}*A_Aussen+A_Anschluesse)*(1-{zr_val})"
                   f"-C{ROW_Z_SANNETTO}+C{ROW_Z_GRUNDSTUECK}")
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = FONT_FORMULA
        cell.number_format = EUR_FMT
    r += 1
SENS_LAST_ROW = r - 1
r += 1
ws.cell(row=r, column=2,
    value="Hinweis: eigenständig für jede Kombination berechnete Tabelle (bewusst kein natives Excel-"
          "Datatable, da TABLE()-Arrayformeln bei openpyxl/LibreOffice-Neuberechnung unzuverlässig sind); "
          "Ergebnis ist identisch zu einer zweidimensionalen What-If-Tabelle.").font = FONT_COMMENT

autofit(ws, {"A": 2, "B": 46, "C": 18, "D": 18, "E": 18})
print("15_Zielpreis OK")

# =====================================================================================
# 16_CO2_Betrieb — Klimaneutralität im Betrieb
# =====================================================================================
ws = wb.create_sheet("16_CO2_Betrieb")
SHEET16 = "16_CO2_Betrieb"
set_title(ws, "16_CO2_Betrieb — Klimaneutralität im Betrieb  [Beim Kopieren: Ziffernpräfix behalten, Rest durch Objekt-ID ersetzen]")

r = 3
section(ws, r, "CO2-Bilanz Heizung/Strom", span=4); r += 1

def corow(label, formula, is_result=False, fmt="#,##0", font_override=None, unit=None):
    global r
    ws.cell(row=r, column=2, value=label).font = f(bold=is_result)
    c = ws.cell(row=r, column=3, value=formula)
    c.number_format = fmt
    c.font = font_override or (FONT_RESULT if is_result else FONT_FORMULA)
    if is_result:
        c.fill = FILL_RESULT
    if unit:
        ws.cell(row=r, column=4, value=unit).font = FONT_COMMENT
    r += 1
    return r - 1

ROW_C_VORHER = corow("CO2-Emissionen vorher",
    f"='{SHEET14}'!$C${ROW_E_VERBRAUCH_ALT}*'{SHEET14}'!$C${ROW_E_CO2_ALT}/1000", unit="kg/a",
    font_override=FONT_LINK)
ROW_C_NETZ = corow("CO2-Emissionen Netzbezug nachher", f"='{SHEET14}'!$C${ROW_E_NETZBEZUG}*A_CO2_Strom/1000",
    unit="kg/a")
ROW_C_GUTSCHRIFT = corow("CO2-Gutschrift PV-Einspeisung",
    f"='{SHEET14}'!$C${ROW_E_PV_UEBERSCHUSS}*A_CO2_PV/1000", unit="kg/a")
ROW_C_NACHHER = corow("→ CO2-Emissionen nachher, netto", f"=C{ROW_C_NETZ}-C{ROW_C_GUTSCHRIFT}", is_result=True)
ROW_C_EINSPARUNG = corow("→ CO2-Einsparung ggü. vorher", f"=C{ROW_C_VORHER}-C{ROW_C_NACHHER}", is_result=True)

r += 1
section(ws, r, "Bilanzielle Deckung durch PV", span=4); r += 1
ROW_C_DECKUNG = corow("Bilanzielle Deckung (PV-Ertrag ÷ Gesamtstrombedarf Haus)",
    f"='{SHEET14}'!$C${ROW_E_PV_ERTRAG}/'{SHEET14}'!$C${ROW_E_STROM_GESAMT}", fmt=PCT_FMT)

HEIZUNG_WP_AKTIV = f'OR({chk_cell("IN_KALK", "A01")}="JA",{chk_cell("IN_KALK", "A02")}="JA")'
ROW_C_FOSSIL_BLOCK = corow("Fossile Heizung bleibt im Zielzustand?",
    f'=IF({HEIZUNG_WP_AKTIV},"Nein - Wärmepumpe aktiv","Ja - keine Wärmepumpe in Kalkulation")', fmt="General")

ROW_C_AMPEL = corow("→ Ampel: bilanzielle Klimaneutralität",
    f'=IF(NOT({HEIZUNG_WP_AKTIV}),"nicht erreichbar (fossile Heizung im Zielzustand)",'
    f'IF(C{ROW_C_DECKUNG}>=1,"gruen: bilanziell klimaneutral",'
    f'IF(C{ROW_C_DECKUNG}>=0.7,"gelb: teilweise gedeckt","rot: geringe Deckung")))',
    fmt="General", is_result=True)

ws.conditional_formatting.add(f"C{ROW_C_AMPEL}",
    FormulaRule(formula=[f'LEFT(C{ROW_C_AMPEL},5)="gruen"'], fill=FILL_RESULT))
ws.conditional_formatting.add(f"C{ROW_C_AMPEL}",
    FormulaRule(formula=[f'LEFT(C{ROW_C_AMPEL},4)="gelb"'], fill=PatternFill("solid", fgColor="FFEB9C")))
ws.conditional_formatting.add(f"C{ROW_C_AMPEL}",
    FormulaRule(formula=[f'OR(LEFT(C{ROW_C_AMPEL},3)="rot",LEFT(C{ROW_C_AMPEL},5)="nicht")'], fill=FILL_WARN))

autofit(ws, {"A": 2, "B": 46, "C": 20, "D": 12})
print("16_CO2_Betrieb OK")

wb.save(OUT)
print("Zwischenspeichern OK:", OUT)

# =====================================================================================
# 00_Objektindex — Übersicht aller angelegten Objekte (Navigations- und Datenquelle für 20_Dashboard)
# =====================================================================================
ws = wb.create_sheet("00_Objektindex", 1)
SHEET_OI = "00_Objektindex"
set_title(ws, "00_Objektindex — Übersicht aller angelegten Objekte")

r = 3
hinweis = ws.cell(row=r, column=2,
    value="Trage hier für jedes angelegte Objekt nur die Objekt-ID ein (der Teil nach dem Ziffernpräfix in den "
          "kopierten Blattnamen, z.B. 10_Rahden1 -> ID = Rahden1). Adresse, Kaufpreis, Wohnfläche und die "
          "Sprunglinks werden automatisch per INDIRECT()-Formel aus dem jeweiligen Objektblatt gezogen — "
          "kein Power Query nötig, kein 'Alle aktualisieren'. Leere Zeilen bleiben leer, bis eine Objekt-ID "
          "eingetragen wird.")
hinweis.font = FONT_COMMENT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
hinweis.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 45
r += 2

header_row(ws, r, ["Objekt-ID", "Adresse", "PLZ", "Ort", "Baujahr", "Kaufpreis €", "Wohnfläche m²",
                    "Zielstandard", "Link 10_<ID>", "Link 11_<ID>", "Link 20_Dashboard"], col=2)
r += 1
OI_FIRST = r
N_OBJEKTE_MAX = 15
OI_ROWS = []
for i in range(N_OBJEKTE_MAX):
    row = r
    OI_ROWS.append(row)
    default_id = ""  # leer lassen: erst nach dem ersten "Blattgruppe kopieren" eine echte Objekt-ID eintragen
    idc = ws.cell(row=row, column=2, value=default_id)
    idc.font = FONT_INPUT; idc.fill = FILL_INPUT
    id_ref = f"$B{row}"

    def mk(colnum, key, fmt=None):
        rownum = OBJ[key]
        cell = ws.cell(row=row, column=colnum,
                        value=f'=IF({id_ref}="","",IFERROR(INDIRECT("\'10_"&{id_ref}&"\'!C{rownum}"),"#REF"))')
        cell.font = FONT_LINK
        if fmt:
            cell.number_format = fmt
        return cell

    mk(3, "Adresse")
    mk(4, "PLZ")
    mk(5, "Ort")
    mk(6, "Baujahr")
    mk(7, "Kaufpreis", fmt=EUR_FMT)
    mk(8, "Wohnflaeche")
    mk(9, "Zielstandard")

    link10 = ws.cell(row=row, column=10,
        value=f'=IF({id_ref}="","",HYPERLINK("#\'10_"&{id_ref}&"\'!A1","10_"&{id_ref}))')
    link10.font = FONT_LINK
    link11 = ws.cell(row=row, column=11,
        value=f'=IF({id_ref}="","",HYPERLINK("#\'11_"&{id_ref}&"\'!A1","11_"&{id_ref}))')
    link11.font = FONT_LINK
    linkdash = ws.cell(row=row, column=12,
        value=f'=IF({id_ref}="","",HYPERLINK("#\'20_Dashboard\'!A1","Dashboard"))')
    linkdash.font = FONT_LINK
    r += 1
OI_LAST = r - 1

ws.cell(row=r + 1, column=2,
    value="Hinweis: Objekt-ID muss exakt mit dem Blattnamen-Suffix übereinstimmen (Groß-/Kleinschreibung "
          "zählt). Bei Tippfehler zeigen die Formeln '#REF' statt echter Werte.").font = FONT_COMMENT

autofit(ws, {"A": 2, "B": 16, "C": 26, "D": 8, "E": 14, "F": 10, "G": 14, "H": 14, "I": 16, "J": 14, "K": 14, "L": 16})
print(f"00_Objektindex OK: Zeilen {OI_FIRST}-{OI_LAST}")

# =====================================================================================
# 20_Dashboard — Objektvergleich (INDIRECT-basiert statt Power Query, siehe 00_Anleitung)
# =====================================================================================
ws = wb.create_sheet("20_Dashboard")
SHEET20 = "20_Dashboard"
set_title(ws, "20_Dashboard — Objektvergleich")

r = 3
hinweis2 = ws.cell(row=r, column=2,
    value="Diese Tabelle übernimmt die Objekt-ID-Liste automatisch aus 00_Objektindex (Spalte B) und zieht "
          "alle Kennzahlen per INDIRECT() aus den zugehörigen 11_<ID>/12_<ID>/14_<ID>/15_<ID>/16_<ID>-Blättern. "
          "Neues Objekt anlegen -> ID auf 00_Objektindex eintragen -> Zeile erscheint hier automatisch, ohne "
          "manuelles Aktualisieren.")
hinweis2.font = FONT_COMMENT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=15)
hinweis2.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 32
r += 2

DASH_HEADERS = ["Objekt-ID", "Adresse", "Baujahr", "Kaufpreis €", "Wohnfläche m²", "€/m² Kaufpreis",
                "Sanierungskosten netto €", "Gesamtinvestition €", "Max. Kaufpreis (Zielpreis) €",
                "Verhandlungsdelta €", "Monatsrate €", "Energiekosten nachher €/a", "CO2-Deckung %",
                "Ampel Gesamt"]
DASH_HEADER_ROW = r
header_row(ws, r, DASH_HEADERS, col=2)
r += 1
DASH_FIRST = r
for i, oi_row in enumerate(OI_ROWS):
    row = r
    id_ref = f"$B{row}"
    ws.cell(row=row, column=2, value=f"='{SHEET_OI}'!B{oi_row}").font = FONT_LINK
    ws.cell(row=row, column=3, value=f"='{SHEET_OI}'!C{oi_row}").font = FONT_LINK
    ws.cell(row=row, column=4, value=f"='{SHEET_OI}'!F{oi_row}").font = FONT_LINK
    kpc = ws.cell(row=row, column=5, value=f"='{SHEET_OI}'!G{oi_row}"); kpc.font = FONT_LINK; kpc.number_format = EUR_FMT
    wfc = ws.cell(row=row, column=6, value=f"='{SHEET_OI}'!H{oi_row}"); wfc.font = FONT_LINK
    epm2 = ws.cell(row=row, column=7,
        value=f'=IF({id_ref}="","",IFERROR(E{row}/F{row},""))')
    epm2.font = FONT_FORMULA; epm2.number_format = EUR_FMT

    def dind(colnum, sheet_prefix, rownum, fmt=EUR_FMT):
        cell = ws.cell(row=row, column=colnum,
            value=f'=IF({id_ref}="","",IFERROR(INDIRECT("\'{sheet_prefix}"&{id_ref}&"\'!C{rownum}"),"#REF"))')
        cell.font = FONT_LINK
        cell.number_format = fmt
        return cell

    dind(8, "11_", ROW_SAN_GESAMT)
    dind(9, "11_", ROW_NETTOINV)
    dind(10, "15_", ROW_Z_MAXGESAMT)
    dind(11, "15_", ROW_Z_DELTA)
    dind(12, "12_", ROW_RATE_MONAT)
    dind(13, "14_", ROW_E_KOSTEN_NACHHER)
    dind(14, "16_", ROW_C_DECKUNG, fmt=PCT_FMT)

    ampel_formula = (
        f'=IF({id_ref}="","",IF(AND(K{row}>=0,N{row}>=0.7,'
        f'L{row}<=A_Einkommen/12*A_Tragfaehigkeit_Quote),"gruen",'
        f'IF(OR(K{row}<0,L{row}>A_Einkommen/12*A_Tragfaehigkeit_Quote),"rot","gelb")))'
    )
    ac = ws.cell(row=row, column=15, value=ampel_formula)
    ac.font = FONT_RESULT
    r += 1
DASH_LAST = r - 1

ws.conditional_formatting.add(f"O{DASH_FIRST}:O{DASH_LAST}",
    FormulaRule(formula=[f'LEFT(O{DASH_FIRST},4)="gruen"'], fill=FILL_RESULT))
ws.conditional_formatting.add(f"O{DASH_FIRST}:O{DASH_LAST}",
    FormulaRule(formula=[f'LEFT(O{DASH_FIRST},4)="gelb"'], fill=PatternFill("solid", fgColor="FFEB9C")))
ws.conditional_formatting.add(f"O{DASH_FIRST}:O{DASH_LAST}",
    FormulaRule(formula=[f'LEFT(O{DASH_FIRST},3)="rot"'], fill=FILL_WARN))

tab_ref = f"B{DASH_HEADER_ROW}:O{DASH_LAST}"
tbl = Table(displayName="TblDashboard", ref=tab_ref)
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

autofit(ws, {"A": 2, "B": 14, "C": 24, "D": 10, "E": 14, "F": 14, "G": 14, "H": 18, "I": 18, "J": 20,
             "K": 16, "L": 14, "M": 18, "N": 12, "O": 14})
ws.freeze_panes = f"C{DASH_FIRST}"
print(f"20_Dashboard OK: Zeilen {DASH_FIRST}-{DASH_LAST}")

wb.save(OUT)
print("Zwischenspeichern OK:", OUT)

# =====================================================================================
# 90_Quellen — Nachvollziehbarkeit der Annahmen
# =====================================================================================
ws = wb.create_sheet("90_Quellen")
set_title(ws, "90_Quellen — Nachvollziehbarkeit der Annahmen")

r = 3
hinweis3 = ws.cell(row=r, column=2,
    value="Eine Zeile je wichtiger/unsicherer Annahme. Ampel-Farbe der Zelle in Spalte 'Verlässlichkeit' "
          "zeigt die Belastbarkeit: grün = belegt/primär, gelb = abgeleitet/berechnet oder Sekundärquelle, "
          "rot = unverifiziert / vor Verwendung zwingend prüfen.")
hinweis3.font = FONT_COMMENT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
hinweis3.alignment = Alignment(wrap_text=True)
r += 2

header_row(ws, r, ["Parameter", "Wert/Bereich", "Quelle", "Abrufdatum", "Verlässlichkeit", "Prüfhinweis vor Verwendung"], col=2)
r += 1
Q_FIRST = r

FILL_GRUEN = FILL_RESULT
FILL_GELB = PatternFill("solid", fgColor="FFEB9C")
FILL_ROT = FILL_WARN

QUELLEN = [
    ("A_Zins_KfW261 — KfW 261 Zinssatz", "3,0% eff. (Planwert)", "kfw.de, Konditionenübersicht",
     "20.08.2026", "rot", "Auf kfw.de zum Recherchezeitpunkt nur Platzhalter sichtbar - vor Kreditantrag beim "
     "Finanzierungspartner den tagesaktuellen Satz verifizieren."),
    ("A_Zins_KfW358 — KfW 358 Ergänzungskredit Zinssatz", "2,5% eff. (Planwert)", "kfw.de", "20.08.2026", "rot",
     "Wie A_Zins_KfW261 unverifiziert - vor Antrag prüfen."),
    ("Worst-Performing-Building-Bonus (WPB)", "Kriterien in Modell nicht separat abgebildet", "BEG-Reform-"
     "Ankündigung 21.07.2026 vs. BAFA-Merkblatt", "20.08.2026", "rot",
     "Merkblatt und Reformankündigung widersprechen sich bei den Schwellenwerten - vor Antragstellung mit "
     "aktuellem Merkblatt und Energieberater abgleichen."),
    ("iSFP-Bonus-Berechnungsweise (Obj_iSFP_JN)", "Modell rechnet mit erhöhtem Fördersatz/Deckel bei iSFP",
     "BAFA-Merkblatt BEG-EM", "20.08.2026", "gelb",
     "Unklar, ob Bonus auf die GESAMTEN förderfähigen Kosten oder nur den Betrag OBERHALB der Standard-"
     "Schwelle gewährt wird - wirtschaftlich relevanter Unterschied, vor Antragstellung klären."),
    ("KNX-Förderfähigkeit (S01-S03 Smart Home)", "Modell behandelt Smart-Home-Maßnahmen konservativ als NICHT "
     "regulär BEG-förderfähig", "BEG-Merkblatt, technische Mindestanforderungen", "20.08.2026", "rot",
     "Technische Mindestanforderungen für eine etwaige Förderfähigkeit einzelner KNX-Komponenten konnten "
     "nicht abschließend recherchiert werden - im Zweifel als ungefördert kalkulieren."),
    ("A_Regio — Regionalfaktor OWL", "1,00 (neutral)", "BKI-Baukosteninformationszentrum (keine frei "
     "zugängliche Regionaltabelle gefunden)", "20.08.2026", "rot",
     "Keine belastbare Quelle für einen Regionalfaktor Ostwestfalen-Lippe gefunden - Sensitivitätstabelle auf "
     "15_Zielpreis zeigt die Bandbreite bei ±10%."),
    ("A_GrESt — Grunderwerbsteuer-Freibetrag NRW", "kein Freibetrag für Selbstnutzer angesetzt", "Ankündigung "
     "Landesregierung NRW zu Öffnungsklausel", "20.08.2026", "gelb",
     "Ein Freibetrag für selbstnutzende Ersterwerber wurde politisch angekündigt, Umsetzungsstatus zum "
     "Recherchezeitpunkt unklar - vor Notartermin beim Finanzamt/Notar erfragen."),
    ("A_Zins_60/80/100 — Bankzinsen nach Beleihungsauslauf", "3,83% / 3,90% / 4,19% eff.", "Interhyp Zins-"
     "chart", "20.08.2026", "gelb", "Tagesaktuelle Marktzinsen, ändern sich laufend - vor Finanzierungszusage "
     "aktuelle Konditionen bei mehreren Anbietern einholen."),
    ("Obj_Bodenrichtwert — Bodenrichtwert je Objekt", "Beispielwert 100 €/m², objektspezifisch einzutragen",
     "BORIS-NRW / Gutachterausschuss (GMB)", "-", "gelb",
     "Für jedes reale Objekt den amtlichen Bodenrichtwert der jeweiligen Zone bei BORIS-NRW nachschlagen und "
     "den Anpassungsfaktor bei stark abweichender Grundstücksgröße prüfen."),
    ("A_Baupreis / A_Baupreis_IH — Baupreissteigerung", "5,0% p.a. / 5,6% p.a.", "Destatis, Baupreisindex "
     "Wohngebäude, Mai 2026", "20.08.2026", "grün", "Amtliche Statistik, regelmäßig neu erhoben - bei älterem "
     "Auszug aktuellste Veröffentlichung auf destatis.de prüfen."),
    ("A_P_Strom/Gas/Oel/Pellet — Energiepreise", "diverse, siehe 01_Annahmen", "BDEW-Strompreisanalyse / "
     "TECSON / DEPV", "20.08.2026", "gelb", "Marktpreise schwanken - vor Betriebskostenkalkulation mit "
     "aktuellem eigenem oder regionalem Tarif abgleichen."),
    ("A_P_Fernwaerme / A_CO2_Fernwaerme", "0,12 €/kWh / 180 g/kWh (grobe Richtwerte)", "keine belastbare "
     "bundesweite Quelle, Fernwärme-Konditionen sind netzbetreiberspezifisch", "20.08.2026", "rot",
     "Fernwärmepreise und CO2-Faktoren unterscheiden sich stark je Netzbetreiber - beim örtlichen Versorger "
     "erfragen, falls ein Objekt Fernwärme nutzt."),
    ("A_CO2_Strom/Gas/Oel/Pellet/PV — CO2-Faktoren", "diverse, siehe 01_Annahmen", "Umweltbundesamt (UBA) "
     "Strommix 2025", "20.08.2026", "grün", "UBA-Werte werden jährlich aktualisiert - bei Bedarf neueste "
     "Fassung prüfen."),
    ("Nutzungsdauern (04_Nutzungsdauern)", "diverse, siehe Blatt", "BTE-Arbeitsblatt 2008 / VDI 2067; "
     "BBSR-Neufassung 2025 nur für WDVS bestätigt aktuell", "20.08.2026", "gelb",
     "Für WDVS ist die BBSR-Neufassung 2025 die aktuellere Quelle; für alle anderen Positionen wurde auf das "
     "ältere BTE-Arbeitsblatt/VDI 2067 zurückgegriffen, da keine neuere Gesamttabelle gefunden wurde."),
    ("A_HH_Strom_Person — Haushaltsstrombedarf/Person", "1.300 kWh/a (Faustwert)", "BDEW-Stromspiegel "
     "(Näherung, keine Skaleneffekt-Korrektur)", "20.08.2026", "gelb",
     "Grobe lineare Näherung - bei großen Haushalten (>4 Personen) tendenziell zu hoch, bei 1-Personen-"
     "Haushalten tendenziell zu niedrig angesetzt."),
]

for name, wert, quelle, datum, ampel, hinweis_txt in QUELLEN:
    ws.cell(row=r, column=2, value=name).font = f()
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=3, value=wert).font = f()
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=4, value=quelle).font = FONT_COMMENT
    ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=5, value=datum).font = FONT_COMMENT
    ampel_cell = ws.cell(row=r, column=6, value={"grün": "belegt/primär", "gelb": "abgeleitet/sekundär",
                                                  "rot": "unverifiziert"}[ampel])
    ampel_cell.fill = {"grün": FILL_GRUEN, "gelb": FILL_GELB, "rot": FILL_ROT}[ampel]
    ampel_cell.font = f(bold=True)
    hc = ws.cell(row=r, column=7, value=hinweis_txt)
    hc.font = f(); hc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 45
    r += 1
Q_LAST = r - 1

autofit(ws, {"A": 2, "B": 34, "C": 26, "D": 34, "E": 12, "F": 16, "G": 46})
ws.freeze_panes = f"B{Q_FIRST}"
print(f"90_Quellen OK: Zeilen {Q_FIRST}-{Q_LAST}")

wb.save(OUT)
print("Zwischenspeichern OK:", OUT)

# =====================================================================================
# 99_Tests — Qualitätssicherung anhand des Referenzobjekts 10_VORLAGE (Default-Werte)
# =====================================================================================
ws = wb.create_sheet("99_Tests")
set_title(ws, "99_Tests — Qualitätssicherung (Referenzobjekt = Default-Werte auf 10_VORLAGE)")

r = 3
hinweis4 = ws.cell(row=r, column=2,
    value="Referenzobjekt: die unveränderten Default-Werte auf 10_VORLAGE (Musterstraße 1, 32312 Lübbecke, "
          "Baujahr 1978, 130 m² Wohnfläche, Kaufpreis 285.000 €, Zielstandard EH 55 EE). Die Soll-Werte wurden "
          "unabhängig von Hand nachgerechnet (siehe Kommentar je Zeile). WICHTIG: Diese Tests gelten nur, "
          "solange 10_VORLAGE nicht verändert wird - für reale Objekte bitte kopierte Blätter (10_<ID> usw.) "
          "verwenden, niemals die VORLAGE selbst überschreiben.")
hinweis4.font = FONT_COMMENT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
hinweis4.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 60
r += 2

header_row(ws, r, ["Prüfpunkt", "Soll (von Hand berechnet)", "Ist (live aus Modell)", "Differenz",
                    "Status", "Herleitung Soll-Wert"], col=2)
r += 1
T_FIRST = r

TESTS = [
    ("1. Kaufnebenkosten-Summe", 318829.50, f"='11_Kalkulation'!C8",
     "285.000×(1+0,065+0,018+0,0357) = 285.000×1,1187 = 318.829,50 € "
     "(GrESt 18.525 + Notar 5.130 + Makler 10.174,50 auf Kaufpreis 285.000 €)"),
    ("2a. Maßnahme H01 (Dach) - Restlebensdauer", -9.5, f"='10_VORLAGE'!P{CHK_ROW_BY_ID['H01']}",
     "Nutzungsdauer 55J × Zustandsfaktor 0,7 (Note 3) - (2026-1978) = 38,5 - 48 = -9,5 Jahre"),
    ("2b. Maßnahme H01 (Dach) - Fälligkeitsjahr", 2026, f"='10_VORLAGE'!Q{CHK_ROW_BY_ID['H01']}",
     "2026 + MAX(0; -9,5) = 2026 (bereits überfällig -> sofort in Kalkulation)"),
    ("2c. Maßnahme H01 (Dach) - Kosten", 43940.00, f"='10_VORLAGE'!J{CHK_ROW_BY_ID['H01']}",
     "Dachfläche 84,5 m² (= Grundfläche 65 m² × Faktor 1,30 Satteldach) × 520 €/m² = 43.940 € "
     "(Preis am 20.08.2026 von 420 auf 520 €/m² angehoben - Nutzer-Praxistest ergab unrealistisch "
     "niedrige Dach-Sanierungskosten; Recherche u.a. Schwäbisch Hall 2026 nennt 400-600€/m² für eine "
     "umfangreiche Dachsanierung, siehe catalog_data.py H01-Kommentar)"),
    ("3. Förderung Weg A gesamt (Effizienzhaus KfW 261)", 21000.00, "='11_Kalkulation'!C23",
     "Tilgungszuschuss 150.000×5% = 7.500 € + Zinsvorteil (3,90%-3,00%)×150.000×10J = 13.500 € = 21.000 € "
     "(unverändert durch H01-Preisanpassung, da Kreditbetrag Basis am 150.000€-Deckel gedeckelt ist)"),
    ("4a. Tilgungsplan Tranche 2, Jahr 1 - Zins €", 15323.42, "='12_Finanzierung'!E65",
     "Restschuld Start 408.624,45 € (gestiegen durch höhere Sanierungssumme nach H01-Preisanpassung, "
     "siehe Test 2c: +8.450€ Rohsumme +253,50€ Fachplanung-Aufschlag [3% von Hülle+Anlagentechnik] "
     "= +8.703,50€, ×1,12 Risikopuffer = +9.748,00€ auf C13, davon -126,75€ höhere Förderung Weg B "
     "[Fachplanung-Zuschuss 50%] = netto +9.621,17€ auf die Nettoinvestition) × 3,75% (Basiszins 3,90% "
     "- Grün-Rabatt 0,15%) = 15.323,42 €"),
    ("4b. Tilgungsplan Tranche 2, Jahr 1 - Tilgung €", 10215.61, "='12_Finanzierung'!F65",
     "Annuität 408.624,45×(3,75%+2,5%)=25.539,03 € - Zins 15.323,42 € = 10.215,61 €"),
    ("4c. Tilgungsplan Tranche 2, Jahr 1 - Restschuld Ende €", 398408.84, "='12_Finanzierung'!H65",
     "Restschuld Start 408.624,45 € - Tilgung 10.215,61 € = 398.408,84 €"),
    ("5. Zielpreis - Max. akzeptabler Kaufpreis gesamt", 149673.05, "='15_Zielpreis'!C15",
     "Neubau 130m²×(2.700+150)×1,00=370.500€; ×1,17 Baunebenkosten +130×270 Außenanlagen +12.000 Anschlüsse "
     "=480.585€ Neubaukosten gesamt; ×(1-20%)=384.468€ Zielwert gebraucht; -309.794,95€ Sanierung netto "
     "(gestiegen durch H01-Preisanpassung, siehe Test 4a) +75.000€ Grundstück (750m²×100€/m²) = 149.673,05 €"),
]

for label, soll, ist_formula, herleitung in TESTS:
    ws.cell(row=r, column=2, value=label).font = f()
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    soll_cell = ws.cell(row=r, column=3, value=soll)
    soll_cell.font = FONT_INPUT; soll_cell.number_format = EUR2_FMT
    ist_cell = ws.cell(row=r, column=4, value=ist_formula)
    ist_cell.font = FONT_LINK; ist_cell.number_format = EUR2_FMT
    diff_cell = ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
    diff_cell.font = FONT_FORMULA; diff_cell.number_format = EUR2_FMT
    status_cell = ws.cell(row=r, column=6, value=f'=IF(ABS(E{r})<0.01,"OK","FEHLER")')
    status_cell.font = f(bold=True)
    her_cell = ws.cell(row=r, column=7, value=herleitung)
    her_cell.font = FONT_COMMENT; her_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 40
    r += 1
T_LAST = r - 1

ws.conditional_formatting.add(f"F{T_FIRST}:F{T_LAST}",
    CellIsRule(operator="equal", formula=['"OK"'], fill=FILL_RESULT))
ws.conditional_formatting.add(f"F{T_FIRST}:F{T_LAST}",
    CellIsRule(operator="equal", formula=['"FEHLER"'], fill=FILL_WARN))

r += 1
ws.cell(row=r, column=2, value="→ Gesamtstatus").font = FONT_RESULT
gesamt_cell = ws.cell(row=r, column=3,
    value=f'=IF(COUNTIF(F{T_FIRST}:F{T_LAST},"FEHLER")=0,"ALLE TESTS OK","FEHLER VORHANDEN - PRÜFEN")')
gesamt_cell.font = FONT_RESULT; gesamt_cell.fill = FILL_RESULT

autofit(ws, {"A": 2, "B": 34, "C": 18, "D": 18, "E": 14, "F": 12, "G": 60})
print(f"99_Tests OK: Zeilen {T_FIRST}-{T_LAST}")

# =====================================================================================
# Abschluss-Politur: Reiterfarben zur optischen Gruppierung, Gitternetzlinien aus
# =====================================================================================
TAB_COLORS = {
    "00_Anleitung": "808080", "00_Objektindex": "808080",
    "01_Annahmen": "1F4E5F", "02_Massnahmenkatalog": "1F4E5F", "03_Foerderprogramme": "1F4E5F",
    "04_Nutzungsdauern": "1F4E5F", "05_Energiekennwerte": "1F4E5F",
    "10_VORLAGE": "375623", "11_Kalkulation": "375623", "12_Finanzierung": "375623",
    "13_Bauzeit": "375623", "14_Betriebskosten": "375623", "15_Zielpreis": "375623",
    "16_CO2_Betrieb": "375623",
    "20_Dashboard": "BF8F00", "90_Quellen": "808080", "99_Tests": "C00000",
}
for sheet_name, color in TAB_COLORS.items():
    wb[sheet_name].sheet_properties.tabColor = color
    wb[sheet_name].sheet_view.showGridLines = False

wb.active = 0
wb.save(OUT)
print("Zwischenspeichern OK:", OUT)
