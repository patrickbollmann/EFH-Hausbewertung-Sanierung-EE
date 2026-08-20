"""
add_object.py — Neues Objekt (Haus) automatisch in Hausbewertung_Altkreis_Luebbecke.xlsx anlegen.

Ersetzt den manuellen "Blattgruppe kopieren + umbenennen"-Schritt aus 00_Anleitung
durch ein Skript. Das manuelle Vorgehen funktioniert weiterhin unverändert (siehe
00_Anleitung in der Datei) - dieses Skript ist eine Abkürzung, kein Ersatz für das
Konzept.

Was das Skript macht:
  1. Kopiert die sieben Objekt-Blätter (10_VORLAGE, 11_Kalkulation, 12_Finanzierung,
     13_Bauzeit, 14_Betriebskosten, 15_Zielpreis, 16_CO2_Betrieb) und benennt die
     Kopien in 10_<ID>, 11_<ID>, ..., 16_<ID> um.
  2. Schreibt alle internen Querverweise zwischen diesen sieben Blättern auf die neuen
     Namen um (Formeln, Diagramm-Datenreihen).
  3. Rekonstruiert, was openpyxl beim Kopieren NICHT automatisch mitnimmt:
     blattlokale benannte Bereiche (Obj_*), Dropdown-Datenüberprüfungen, bedingte
     Formatierung, Diagramme (Kuchendiagramm auf 11_<ID>, Balkendiagramme auf
     13_<ID>/14_<ID>), Fensterfixierung und Autofilter.
  4. Trägt die neue Objekt-ID in die nächste freie Zeile auf 00_Objektindex ein -
     20_Dashboard zieht die Kennzahlen danach automatisch per INDIRECT() nach.
  5. Legt vor dem Speichern ein Backup der Originaldatei an.
  6. Ruft optional die LibreOffice-Neuberechnung (recalc.py aus dem xlsx-Skill) auf,
     falls verfügbar, und meldet das Ergebnis.

Nutzung:
  python3 add_object.py Rahden1
  python3 add_object.py Rahden1 --adresse "Bahnhofstr. 12" --plz 32369 --ort Rahden \
      --baujahr 1985 --kaufpreis 265000 --wohnflaeche 145 \
      --grundstuecksflaeche 600 --bodenrichtwert 95

  python3 add_object.py Rahden1 --file /pfad/zu/Hausbewertung_Altkreis_Luebbecke.xlsx

  # Alle Stammdaten + Checklisten-Overrides aus einer JSON-Datei (z.B. von einer KI aus
  # einem Exposé extrahiert, siehe docs/Objekt_aus_Expose_anlegen.md):
  python3 add_object.py Rahden1 --data rahden1.json

Voraussetzung: openpyxl (pip install openpyxl --break-system-packages, falls fehlend).
"""
import argparse
import copy
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

from catalog_data import CATALOG

TEMPLATE_SHEETS = [
    "10_VORLAGE",
    "11_Kalkulation",
    "12_Finanzierung",
    "13_Bauzeit",
    "14_Betriebskosten",
    "15_Zielpreis",
    "16_CO2_Betrieb",
]

# Label -> (Blatt, Zeilentext in Spalte B) für optionale Startwerte per CLI.
FIELD_LABELS = {
    "adresse": "Adresse",
    "plz": "PLZ",
    "ort": "Ort",
    "baujahr": "Baujahr",
    "kaufpreis": "Kaufpreis (Angebot)",
    "wohnflaeche": "Wohnfläche",
    "grundstuecksflaeche": "Grundstücksfläche",
    "bodenrichtwert": "Bodenrichtwert",
}

# argparse liefert CLI-Werte immer als str - hier wird festgelegt, in welchen
# Python-Typ jedes Feld vor dem Schreiben in die Zelle konvertiert werden muss.
# Ohne diese Konvertierung landet z.B. der Kaufpreis als TEXT statt als Zahl in
# der Zelle: SUM() ignoriert Textzellen stillschweigend (0 statt Fehler), während
# eine direkte Multiplikation Text noch automatisch in eine Zahl umwandelt - das
# Resultat wären falsche, aber unauffällige Zwischensummen ein paar Zeilen weiter.
FIELD_TYPES = {
    "adresse": str,
    "plz": int,
    "ort": str,
    "baujahr": int,
    "kaufpreis": float,
    "wohnflaeche": float,
    "grundstuecksflaeche": float,
    "bodenrichtwert": float,
}

# Vollständige Feldliste für --data JSON (alle 28 obj_field()-Einträge auf 10_VORLAGE).
# key -> (Label in Spalte B, Python-Typ, erlaubte Werte oder None für Freitext/Zahl).
# Muss synchron zu den obj_field(...)-Aufrufen in build_workbook.py gehalten werden.
STAMMDATEN_FIELDS = {
    "adresse": ("Adresse", str, None),
    "plz": ("PLZ", int, None),
    "ort": ("Ort", str, None),
    "baujahr": ("Baujahr", int, None),
    "kaufpreis": ("Kaufpreis (Angebot)", float, None),
    "wohnflaeche": ("Wohnfläche", float, None),
    "grundstuecksflaeche": ("Grundstücksfläche", float, None),
    "bodenrichtwert": ("Bodenrichtwert", float, None),
    "anpassungsfaktor": ("Anpassungsfaktor Grundstücksgröße", float, None),
    "geschosse": ("Geschosse", int, None),
    "keller_typ": ("Keller", str, ["Kein Keller", "Teilkeller", "Vollkeller"]),
    "dachform": ("Dachform", str, ["Satteldach", "Flachdach", "Walmdach"]),
    "dach_eternit": ("Dacheindeckung Eternit/Asbestverdacht", str, ["Ja", "Nein"]),
    "energieausweis": ("Energieausweis-Wert (0 = kein Ausweis)", float, None),
    "heizung_alt": ("Heizungsart aktuell", str,
                     ["Gasheizung", "Ölheizung", "Nachtspeicher", "Fernwärme", "Wärmepumpe", "Sonstige"]),
    "anz_nachtspeicher": ("Anzahl Nachtspeicheröfen", int, None),
    "feuchte_keller": ("Feuchteschaden im Keller", str, ["Ja", "Nein"]),
    "zielstandard": ("Zielstandard", str, ["EH 55 EE", "EH 40 EE", "Nur Einzelmaßnahmen"]),
    "isfp_ja_nein": ("iSFP-Bonus nutzen (Energieberater beauftragt)", str, ["Ja", "Nein"]),
    "foerderweg_wahl": ("Förderweg-Wahl", str,
                        ["Automatik (Empfehlung)", "Weg A (Effizienzhaus)", "Weg B (Einzelmaßnahmen)"]),
    "makler_ja_nein": ("Makler beteiligt", str, ["Ja", "Nein"]),
    "anz_baeder": ("Anzahl Bäder", int, None),
    "personen_override": ("Personen im Haushalt (0=Default aus Annahmen)", int, None),
    "sanierungsdauer_monate": ("Sanierungsdauer", int, None),
    "einzug_nach_sanierung": ("Einzug erst nach Sanierung", str, ["Ja", "Nein"]),
    "parallelmiete_monat": ("Miete während Sanierung", float, None),
    "pv_kwp_override": ("PV Zielgröße (0 = automatisch aus Dachfläche)", float, None),
    "pv_ausrichtung": ("PV Ausrichtung", str, ["Süd", "Ost-West"]),
}

# Checkliste (pro Maßnahme aus catalog_data.CATALOG eine Zeile): key -> exakter
# Spaltenkopf-Text aus CHK_HEADERS in build_workbook.py. Spaltenposition wird zur
# Laufzeit per Header-Textsuche ermittelt (nicht hart codiert), damit ein späteres
# Verschieben von Spalten in build_workbook.py hier nicht zu falschen Werten führt.
CHK_ID_HEADER = "ID"
CHK_FIELD_HEADERS = {
    "menge_override": "Menge\nOverride",
    "baujahr": "Baujahr/\nletzte Erneuerung",
    "zustand": "Zustands-\nnote (1-5)",
    "override_zeitpunkt": "Override\nZeitpunkt",
    "eigenleistung_pct": "Eigenleistung\n%",
    "kommentar": "Kommentar",
}
OVERRIDE_ZEITPUNKT_CHOICES = ["Automatik", "Sofort", "In Horizont", "Später", "Entfällt"]
ZUSTAND_CHOICES = [1, 2, 3, 4, 5]
VALID_CATALOG_IDS = {item["id"] for item in CATALOG}

FORBIDDEN_CHARS = set('/\\?*[]:')
MAX_ID_LEN = 31 - len("16_")  # längstes Präfix ist "16_", Excel-Limit 31 Zeichen/Blattname


def fail(msg):
    print(f"FEHLER: {msg}", file=sys.stderr)
    sys.exit(1)


def validate_id(obj_id, wb):
    if not obj_id:
        fail("Objekt-ID darf nicht leer sein.")
    if any(ch in FORBIDDEN_CHARS for ch in obj_id):
        fail(f"Objekt-ID '{obj_id}' enthält ein bei Excel-Blattnamen verbotenes Zeichen ({FORBIDDEN_CHARS}).")
    if len(obj_id) > MAX_ID_LEN:
        fail(f"Objekt-ID '{obj_id}' ist zu lang ({len(obj_id)} Zeichen). Maximal {MAX_ID_LEN} Zeichen erlaubt, "
             f"da Excel Blattnamen auf 31 Zeichen begrenzt und das längste Präfix '16_' ist.")
    for prefix in ("10_", "11_", "12_", "13_", "14_", "15_", "16_"):
        candidate = f"{prefix}{obj_id}"
        if candidate in wb.sheetnames:
            fail(f"Blatt '{candidate}' existiert bereits. Objekt-ID '{obj_id}' ist schon vergeben "
                 f"oder eine gleichnamige Kopie liegt bereits in der Datei.")


def find_row_by_label(ws, label_text, col=2, max_row=200):
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip() == label_text:
            return r
    return None


def rewrite_formula_sheet_refs(formula, rename_map):
    """Ersetzt Verweise auf die alten Vorlagen-Blattnamen (quoted & unquoted) durch die neuen."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    result = formula
    for old_name, new_name in rename_map.items():
        # gequotete Form: 'Alt_Name'!
        result = re.sub(r"'" + re.escape(old_name) + r"'!", f"'{new_name}'!", result)
        # ungequotete Form: Alt_Name! (nur ganzes Token, kein Teiltreffer)
        result = re.sub(r"(?<![\w'])" + re.escape(old_name) + r"!", f"{new_name}!", result)
    return result


def rewrite_chart_refs(chart, rename_map):
    """Läuft rekursiv durch die Formel-Strings (.f-Attribute) eines Diagramm-Objekts
    und ersetzt Blattnamen-Verweise. Diagramme referenzieren Zellen immer über
    numRef.f / strRef.f Strings wie 'Blattname!$C$1:$C$9'."""
    def rewrite_ref(ref_obj):
        if ref_obj is None:
            return
        f = getattr(ref_obj, "f", None)
        if isinstance(f, str):
            new_f = f
            for old_name, new_name in rename_map.items():
                new_f = re.sub(r"'" + re.escape(old_name) + r"'!", f"'{new_name}'!", new_f)
                new_f = re.sub(r"(?<![\w'])" + re.escape(old_name) + r"!", f"{new_name}!", new_f)
            ref_obj.f = new_f

    for ser in getattr(chart, "series", []):
        val = getattr(ser, "val", None)
        if val is not None:
            rewrite_ref(getattr(val, "numRef", None))
        cat = getattr(ser, "cat", None)
        if cat is not None:
            rewrite_ref(getattr(cat, "numRef", None))
            rewrite_ref(getattr(cat, "strRef", None))
        tx = getattr(ser, "tx", None)
        if tx is not None:
            rewrite_ref(getattr(tx, "strRef", None))


def copy_object_group(wb, obj_id):
    rename_map = {old: f"{old.split('_')[0]}_{obj_id}" for old in TEMPLATE_SHEETS}
    new_sheets = {}

    # 1. Blätter kopieren + umbenennen (Reihenfolge wichtig für spätere Referenzsuche)
    for old_name in TEMPLATE_SHEETS:
        src = wb[old_name]
        new_ws = wb.copy_worksheet(src)
        new_ws.title = rename_map[old_name]
        new_ws.sheet_properties.tabColor = copy.copy(src.sheet_properties.tabColor)
        new_ws.sheet_view.showGridLines = False
        new_sheets[old_name] = new_ws
        print(f"  kopiert: {old_name} -> {new_ws.title}")

    # 2. Formeln umschreiben (Querverweise zwischen den 7 Blättern)
    for old_name, new_ws in new_sheets.items():
        for row in new_ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    rewritten = rewrite_formula_sheet_refs(cell.value, rename_map)
                    if rewritten != cell.value:
                        cell.value = rewritten

    # 3. Fensterfixierung + Autofilter (werden von copy_worksheet nicht übernommen)
    for old_name, new_ws in new_sheets.items():
        src = wb[old_name]
        new_ws.freeze_panes = src.freeze_panes
        if src.auto_filter and src.auto_filter.ref:
            new_ws.auto_filter.ref = src.auto_filter.ref

    # 4. Blattlokale benannte Bereiche (Obj_*) rekonstruieren - nur auf 10_<ID> relevant
    src10 = wb["10_VORLAGE"]
    new10 = new_sheets["10_VORLAGE"]
    for name, defn in src10.defined_names.items():
        new_attr = rewrite_formula_sheet_refs("=" + defn.attr_text, rename_map)[1:]
        new_defn = DefinedName(name, attr_text=new_attr)
        new10.defined_names[name] = new_defn
    print(f"  {len(src10.defined_names)} lokale Namen (Obj_*) auf 10_{obj_id} rekonstruiert")

    # 5. Datenüberprüfungen (Dropdowns) - Formeln sind literale Listen ohne Blattbezug,
    #    einfach unverändert übernehmen.
    total_dv = 0
    for old_name, new_ws in new_sheets.items():
        src = wb[old_name]
        for dv in src.data_validations.dataValidation:
            new_ws.add_data_validation(copy.copy(dv))
            total_dv += 1
    print(f"  {total_dv} Dropdown-Validierungen übernommen")

    # 6. Bedingte Formatierung (Formeln sind blattintern relativ, unverändert übernehmen)
    total_cf = 0
    for old_name, new_ws in new_sheets.items():
        src = wb[old_name]
        for cf_range in src.conditional_formatting:
            for rule in cf_range.rules:
                new_ws.conditional_formatting.add(str(cf_range.sqref), copy.copy(rule))
                total_cf += 1
    print(f"  {total_cf} Regeln für bedingte Formatierung übernommen")

    # 7. Diagramme klonen (Kuchendiagramm 11_<ID>, Balkendiagramme 13_<ID>/14_<ID>)
    total_charts = 0
    for old_name, new_ws in new_sheets.items():
        src = wb[old_name]
        for chart in src._charts:
            new_chart = copy.deepcopy(chart)
            rewrite_chart_refs(new_chart, rename_map)
            new_ws.add_chart(new_chart)
            total_charts += 1
    print(f"  {total_charts} Diagramme geklont")

    return new_sheets, rename_map


def reorder_sheets(wb, new_sheet_titles):
    """Verschiebt die neu erzeugten Blätter direkt vor 20_Dashboard (falls vorhanden),
    sonst ans Ende - rein kosmetisch, hat keinen Einfluss auf Formeln."""
    names = wb.sheetnames
    if "20_Dashboard" not in names:
        return
    sheets_by_name = {ws.title: ws for ws in wb._sheets}
    remaining = [ws for ws in wb._sheets if ws.title not in new_sheet_titles]
    insert_at = next(i for i, ws in enumerate(remaining) if ws.title == "20_Dashboard")
    new_ws_objs = [sheets_by_name[t] for t in new_sheet_titles]
    wb._sheets = remaining[:insert_at] + new_ws_objs + remaining[insert_at:]


def fill_initial_values(wb, obj_id, args):
    ws10 = wb[f"10_{obj_id}"]
    for arg_name, label in FIELD_LABELS.items():
        raw_value = getattr(args, arg_name)
        if raw_value is None:
            continue
        target_type = FIELD_TYPES[arg_name]
        try:
            value = target_type(raw_value.replace(",", ".")) if target_type is float else target_type(raw_value)
        except ValueError:
            fail(f"--{arg_name} erwartet einen Wert vom Typ {target_type.__name__}, erhalten: '{raw_value}'")
        row = find_row_by_label(ws10, label)
        if row is None:
            print(f"  WARNUNG: Zeile mit Label '{label}' auf 10_{obj_id} nicht gefunden, "
                  f"überspringe --{arg_name}.")
            continue
        ws10.cell(row=row, column=3, value=value)
        print(f"  gesetzt: {label} = {value!r} ({target_type.__name__})")


def apply_stammdaten(wb, obj_id, stammdaten):
    """Setzt Stammdaten-Felder aus einem --data JSON-Dict (siehe STAMMDATEN_FIELDS).
    Wird VOR fill_initial_values() (CLI-Flags) aufgerufen, sodass ein zusätzlich
    angegebenes CLI-Flag einen JSON-Wert gezielt überschreiben kann."""
    if not isinstance(stammdaten, dict):
        fail("--data JSON: 'stammdaten' muss ein Objekt (JSON-Dict) sein.")
    ws10 = wb[f"10_{obj_id}"]
    unknown = set(stammdaten) - set(STAMMDATEN_FIELDS)
    if unknown:
        fail(f"--data JSON: unbekannte Stammdaten-Felder {sorted(unknown)}. "
             f"Gültige Felder: {', '.join(sorted(STAMMDATEN_FIELDS))}")
    for key, raw_value in stammdaten.items():
        if raw_value is None:
            continue
        label, target_type, choices = STAMMDATEN_FIELDS[key]
        if choices is not None:
            value = str(raw_value)
            if value not in choices:
                fail(f"Stammdaten-Feld '{key}': Wert {value!r} nicht erlaubt. Erlaubt: {choices}")
        else:
            try:
                if target_type is float and isinstance(raw_value, str):
                    value = float(raw_value.replace(",", "."))
                else:
                    value = target_type(raw_value)
            except (TypeError, ValueError):
                fail(f"Stammdaten-Feld '{key}' erwartet Typ {target_type.__name__}, erhalten: {raw_value!r}")
        row = find_row_by_label(ws10, label)
        if row is None:
            print(f"  WARNUNG: Zeile mit Label '{label}' auf 10_{obj_id} nicht gefunden, "
                  f"überspringe Stammdaten-Feld '{key}'.")
            continue
        ws10.cell(row=row, column=3, value=value)
        print(f"  gesetzt (JSON): {label} = {value!r} ({type(value).__name__})")


def find_checkliste_layout(ws10):
    """Findet die Kopfzeile der Sanierungs-Checkliste (Spalte B = 'ID') per Textsuche
    und liefert (col_map, id_to_row) - Spalten/Zeilen werden nie hart codiert, damit
    ein späteres Verschieben von Spalten in build_workbook.py hier nicht zu falschen
    Werten führt, sondern zu einem klaren Fehler."""
    header_row = None
    for r in range(1, 80):
        v = ws10.cell(row=r, column=2).value
        if v is not None and str(v).strip() == CHK_ID_HEADER:
            header_row = r
            break
    if header_row is None:
        fail(f"Kopfzeile der Sanierungs-Checkliste (Spalte B = '{CHK_ID_HEADER}') "
             f"auf 10_{ws10.title.split('_', 1)[-1]} nicht gefunden.")

    header_texts = {}
    for c in range(1, 30):
        v = ws10.cell(row=header_row, column=c).value
        if v is not None:
            header_texts[str(v).strip()] = c

    col_map = {}
    for field, header_text in CHK_FIELD_HEADERS.items():
        if header_text not in header_texts:
            fail(f"Spaltenkopf {header_text!r} für Checklisten-Feld '{field}' nicht gefunden - "
                 f"Layout der Checkliste hat sich vermutlich geändert (CHK_HEADERS in build_workbook.py "
                 f"und CHK_FIELD_HEADERS in add_object.py müssen synchron bleiben).")
        col_map[field] = header_texts[header_text]

    id_col = header_texts.get(CHK_ID_HEADER, 2)
    id_to_row = {}
    r = header_row + 1
    while True:
        v = ws10.cell(row=r, column=id_col).value
        if v is None or str(v).strip() == "":
            break
        id_to_row[str(v).strip()] = r
        r += 1
    return col_map, id_to_row


def apply_checkliste_entries(wb, obj_id, checkliste):
    """Setzt pro Maßnahme (Katalog-ID) Overrides aus einem --data JSON. Jeder Eintrag
    kann eine Teilmenge von menge_override/baujahr/zustand/override_zeitpunkt/
    eigenleistung_pct/kommentar enthalten - alles Felder, die auch beim manuellen
    Ausfüllen der Checkliste editierbar sind."""
    if not isinstance(checkliste, list):
        fail("--data JSON: 'checkliste' muss eine Liste von Objekten sein.")
    ws10 = wb[f"10_{obj_id}"]
    col_map, id_to_row = find_checkliste_layout(ws10)

    for entry in checkliste:
        if not isinstance(entry, dict) or "id" not in entry:
            fail(f"--data JSON: jeder Checklisten-Eintrag braucht ein 'id'-Feld, erhalten: {entry!r}")
        cid = str(entry["id"]).strip()
        if cid not in VALID_CATALOG_IDS:
            fail(f"Checklisten-Eintrag: '{cid}' ist keine gültige Maßnahmen-ID. "
                 f"Gültige IDs stehen in 02_Massnahmenkatalog bzw. build/catalog_data.py "
                 f"(z.B. H01-H12 Hülle, A01-A08 Anlagentechnik, R01-R06 Risiko/Schadstoff, "
                 f"K01-K04 Klimaneutralität, S01-S03 Smart Home, E01-E10 Elektro/Innenausbau, "
                 f"G01-G04 Außenanlagen, W01-W04 weiche Kosten).")
        if cid not in id_to_row:
            fail(f"Maßnahme '{cid}' ist im Katalog bekannt, aber keine Zeile dafür auf "
                 f"10_{obj_id} gefunden (sollte nicht passieren, bitte Datei prüfen).")
        row = id_to_row[cid]
        set_fields = []

        if entry.get("menge_override") is not None:
            try:
                val = float(entry["menge_override"])
            except (TypeError, ValueError):
                fail(f"{cid}: menge_override muss eine Zahl sein, erhalten: {entry['menge_override']!r}")
            ws10.cell(row=row, column=col_map["menge_override"], value=val)
            set_fields.append(f"menge_override={val}")

        if entry.get("baujahr") is not None:
            try:
                val = int(entry["baujahr"])
            except (TypeError, ValueError):
                fail(f"{cid}: baujahr muss eine Ganzzahl sein, erhalten: {entry['baujahr']!r}")
            ws10.cell(row=row, column=col_map["baujahr"], value=val)
            set_fields.append(f"baujahr={val}")

        if entry.get("zustand") is not None:
            try:
                val = int(entry["zustand"])
            except (TypeError, ValueError):
                fail(f"{cid}: zustand muss eine Ganzzahl 1-5 sein, erhalten: {entry['zustand']!r}")
            if val not in ZUSTAND_CHOICES:
                fail(f"{cid}: zustand muss zwischen 1 und 5 liegen, erhalten: {val}")
            ws10.cell(row=row, column=col_map["zustand"], value=val)
            set_fields.append(f"zustand={val}")

        if entry.get("override_zeitpunkt") is not None:
            val = str(entry["override_zeitpunkt"])
            if val not in OVERRIDE_ZEITPUNKT_CHOICES:
                fail(f"{cid}: override_zeitpunkt muss einer von {OVERRIDE_ZEITPUNKT_CHOICES} sein, "
                     f"erhalten: {val!r}")
            ws10.cell(row=row, column=col_map["override_zeitpunkt"], value=val)
            set_fields.append(f"override_zeitpunkt={val}")

        if entry.get("eigenleistung_pct") is not None:
            try:
                val = float(entry["eigenleistung_pct"])
            except (TypeError, ValueError):
                fail(f"{cid}: eigenleistung_pct muss eine Zahl zwischen 0 und 1 sein, "
                     f"erhalten: {entry['eigenleistung_pct']!r}")
            if not (0 <= val <= 1):
                fail(f"{cid}: eigenleistung_pct muss zwischen 0 und 1 liegen (0.2 = 20%), erhalten: {val}")
            ws10.cell(row=row, column=col_map["eigenleistung_pct"], value=val)
            set_fields.append(f"eigenleistung_pct={val}")

        if entry.get("kommentar") is not None:
            val = str(entry["kommentar"])
            ws10.cell(row=row, column=col_map["kommentar"], value=val)
            set_fields.append("kommentar gesetzt")

        if set_fields:
            print(f"  Checkliste {cid} (Zeile {row}): {', '.join(set_fields)}")
        else:
            print(f"  Checkliste {cid} (Zeile {row}): keine Felder im JSON-Eintrag gesetzt, übersprungen.")


def add_to_objektindex(wb, obj_id):
    if "00_Objektindex" not in wb.sheetnames:
        print("  WARNUNG: Blatt 00_Objektindex nicht gefunden - Objekt-ID nicht eingetragen.")
        return
    ws = wb["00_Objektindex"]
    header_row = find_row_by_label(ws, "Objekt-ID")
    if header_row is None:
        print("  WARNUNG: Kopfzeile 'Objekt-ID' auf 00_Objektindex nicht gefunden.")
        return
    r = header_row + 1
    while ws.cell(row=r, column=2).value not in (None, ""):
        existing = str(ws.cell(row=r, column=2).value).strip()
        if existing == obj_id:
            print(f"  Objekt-ID '{obj_id}' steht bereits in Zeile {r} von 00_Objektindex.")
            return
        r += 1
    ws.cell(row=r, column=2, value=obj_id)
    print(f"  Objekt-ID '{obj_id}' in 00_Objektindex, Zeile {r} eingetragen")


def run_recalc(path):
    recalc_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "recalc.py")
    if not os.path.exists(recalc_script):
        print("\nHinweis: build/recalc.py nicht gefunden - bitte die Datei einmal in "
              "Excel/LibreOffice öffnen und speichern, damit alle Formeln berechnete "
              "Werte bekommen.")
        return
    print("\nRechne mit LibreOffice neu (recalc.py) ...")
    result = subprocess.run(
        [sys.executable, recalc_script, path, "150"],
        capture_output=True, text=True,
    )
    print(result.stdout.strip())
    if result.returncode != 0:
        print("WARNUNG: recalc.py meldet einen Fehler (siehe oben, z.B. weil LibreOffice "
              "nicht installiert ist - dann die Datei einmal manuell in Excel/LibreOffice "
              "öffnen und speichern).", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("objekt_id", help="Kurze Objekt-ID ohne Sonderzeichen, z.B. Rahden1")
    parser.add_argument("--file", default=None,
                         help="Pfad zur Excel-Datei (Default: ../Hausbewertung_Altkreis_Luebbecke.xlsx "
                              "relativ zu diesem Skript)")
    parser.add_argument("--no-backup", action="store_true", help="Kein Backup vor dem Überschreiben anlegen")
    parser.add_argument("--no-recalc", action="store_true", help="LibreOffice-Neuberechnung überspringen")
    parser.add_argument("--data", default=None,
                         help="Pfad zu einer JSON-Datei mit 'stammdaten' (Objekt: Feldname -> Wert, siehe "
                              "STAMMDATEN_FIELDS) und/oder 'checkliste' (Liste von Maßnahmen-Overrides). "
                              "Siehe docs/Objekt_aus_Expose_anlegen.md. CLI-Flags wie --kaufpreis überschreiben "
                              "bei Überschneidung den JSON-Wert.")
    for arg_name in FIELD_LABELS:
        parser.add_argument(f"--{arg_name}", default=None, help=f"Startwert für '{FIELD_LABELS[arg_name]}'")
    args = parser.parse_args()

    default_path = os.path.join(os.path.dirname(__file__), "..", "Hausbewertung_Altkreis_Luebbecke.xlsx")
    path = os.path.abspath(args.file or default_path)
    if not os.path.exists(path):
        fail(f"Datei nicht gefunden: {path}")

    data = None
    if args.data:
        data_path = os.path.abspath(args.data)
        if not os.path.exists(data_path):
            fail(f"--data Datei nicht gefunden: {data_path}")
        with open(data_path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError as e:
                fail(f"--data Datei ist kein gültiges JSON: {e}")
        if not isinstance(data, dict):
            fail("--data JSON: oberste Ebene muss ein Objekt mit 'stammdaten' und/oder 'checkliste' sein.")
        unknown_top = set(data) - {"stammdaten", "checkliste"}
        if unknown_top:
            fail(f"--data JSON: unbekannte Top-Level-Schlüssel {sorted(unknown_top)} "
                 f"(erlaubt: 'stammdaten', 'checkliste').")

    print(f"Öffne {path} ...")
    wb = load_workbook(path)
    validate_id(args.objekt_id, wb)

    if not args.no_backup:
        backup_path = f"{path}.bak-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        shutil.copy2(path, backup_path)
        print(f"Backup angelegt: {backup_path}")

    print(f"\nLege Objekt '{args.objekt_id}' an ...")
    new_sheets, rename_map = copy_object_group(wb, args.objekt_id)
    reorder_sheets(wb, list(rename_map.values()))

    if data and data.get("stammdaten"):
        print("\nSetze Stammdaten aus --data JSON ...")
        apply_stammdaten(wb, args.objekt_id, data["stammdaten"])

    fill_initial_values(wb, args.objekt_id, args)

    if data and data.get("checkliste"):
        print("\nSetze Checklisten-Overrides aus --data JSON ...")
        apply_checkliste_entries(wb, args.objekt_id, data["checkliste"])

    add_to_objektindex(wb, args.objekt_id)

    wb.save(path)
    print(f"\nGespeichert: {path}")

    if not args.no_recalc:
        run_recalc(path)

    print(f"\nFertig. Neue Blätter: {', '.join(rename_map.values())}")
    print(f"Auf 10_{args.objekt_id} die restlichen Objektdaten (gelbe Zellen) und die "
          f"Sanierungs-Checkliste ausfüllen. 20_Dashboard zeigt das Objekt automatisch an, "
          f"sobald die Objekt-ID auf 00_Objektindex steht.")


if __name__ == "__main__":
    main()
