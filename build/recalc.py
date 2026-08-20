# -*- coding: utf-8 -*-
"""
recalc.py — berechnet alle Formeln einer .xlsx-Datei per LibreOffice headless neu.

Warum das nötig ist: openpyxl schreibt Formeln nur als Text, ohne berechneten Wert
("<v>"-Cache leer). Excel/LibreOffice berechnen das automatisch beim Öffnen - aber
wer die Datei nur mit openpyxl (data_only=True) ausliest, ohne sie vorher einmal in
einem echten Tabellenkalkulationsprogramm geöffnet zu haben, sieht überall `None`
statt der eigentlichen Werte. Dieses Skript automatisiert genau dieses "einmal
öffnen und speichern" per Kommandozeile.

Voraussetzung: LibreOffice muss installiert sein (Kommando `soffice` auf dem PATH).
  - Debian/Ubuntu: apt install libreoffice-calc
  - macOS: brew install --cask libreoffice
  - Windows: https://www.libreoffice.org/download/

Nutzung:
  python3 recalc.py <pfad-zur-datei.xlsx> [timeout-sekunden]

Gibt ein JSON-Ergebnis auf stdout aus, z.B.:
  {"status": "success", "total_errors": 0, "error_summary": {}, "total_formulas": 2035}

Exit-Code 0 bei Erfolg (auch wenn total_errors > 0 - das Ergebnis wird trotzdem
ausgegeben, damit der Aufrufer selbst entscheiden kann), Exit-Code 1 wenn
LibreOffice nicht gefunden wurde oder der Konvertierungsvorgang fehlgeschlagen ist.
"""
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

ERROR_RE = re.compile(r"^#(REF|NAME\??|VALUE!|DIV/0!|N/A|NULL!|NUM!)$")


def find_soffice():
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def run_recalc(path, timeout=150):
    soffice = find_soffice()
    if soffice is None:
        return {
            "status": "error",
            "message": (
                "LibreOffice (Kommando 'soffice'/'libreoffice') wurde nicht auf dem PATH "
                "gefunden. Bitte installieren (siehe Kopfkommentar dieses Skripts) oder "
                "die Datei einmal manuell in Excel/LibreOffice öffnen und speichern."
            ),
        }

    path = os.path.abspath(path)
    if not os.path.exists(path):
        return {"status": "error", "message": f"Datei nicht gefunden: {path}"}

    with tempfile.TemporaryDirectory() as tmpdir:
        cmd = [soffice, "--headless", "--norestore", "--convert-to", "xlsx",
               "--outdir", tmpdir, path]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            return {"status": "error",
                    "message": f"LibreOffice hat nach {timeout}s nicht reagiert (Timeout)."}

        converted = os.path.join(tmpdir, os.path.basename(path))
        if result.returncode != 0 or not os.path.exists(converted):
            return {
                "status": "error",
                "message": "LibreOffice-Konvertierung fehlgeschlagen.",
                "stdout": result.stdout.strip(),
                "stderr": result.stderr.strip(),
            }
        shutil.copy2(converted, path)

    return scan_errors(path)


def scan_errors(path):
    wb_formulas = load_workbook(path, data_only=False)
    wb_values = load_workbook(path, data_only=True)

    total_formulas = 0
    error_summary = {}
    for ws in wb_formulas.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1

    for ws in wb_values.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and ERROR_RE.match(v):
                    error_summary[v] = error_summary.get(v, 0) + 1

    wb_formulas.close()
    wb_values.close()

    total_errors = sum(error_summary.values())
    return {
        "status": "success",
        "total_errors": total_errors,
        "error_summary": error_summary,
        "total_formulas": total_formulas,
    }


def main():
    if len(sys.argv) < 2:
        print("Nutzung: python3 recalc.py <pfad-zur-datei.xlsx> [timeout-sekunden]", file=sys.stderr)
        sys.exit(1)
    path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 150

    result = run_recalc(path, timeout)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    sys.exit(0 if result.get("status") == "success" else 1)


if __name__ == "__main__":
    main()
